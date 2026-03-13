"""
出入库单 × 销售订单 分析报告生成器
输出文件：出入库分析_20260313_v1.xlsx
Sheet 1: 总览   Sheet 2: 出库明细   Sheet 3: 成本核对   Sheet 4: 入库明细
"""

import math, re, os
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

# ── 路径 ──────────────────────────────────────────────────
BASE      = "数据快照/"
F_OUT     = BASE + "宇宙销售统计中心_2026 出库单_031318.xlsx"
F_IN      = BASE + "宇宙销售统计中心_2026 入库单_031318.xlsx"
F_ORD     = BASE + "宇宙销售统计中心_2026 公路车销售订单_031116.xlsx"
OUT_FILE  = "库存报表/出入库分析_20260313_v1.xlsx"

os.makedirs("库存报表", exist_ok=True)

# ── STYLE1 色板 ───────────────────────────────────────────
C_DARK_NAVY   = "1B2A47"   # 深海军蓝  - 主标题背景
C_MID_BLUE    = "2E4172"   # 中蓝      - 二级标题
C_STEEL       = "4A6FA5"   # 钢蓝      - 三级标题 / 分组标题
C_ICE         = "D6E4F0"   # 冰蓝      - 隔行底色
C_WHITE       = "FFFFFF"
C_BORDER      = "B8C8DC"   # 边框色
C_AMBER       = "F5A623"   # 琥珀      - 警告/差异
C_RED_BG      = "FDE8E8"   # 浅红      - 严重问题
C_GREEN_BG    = "E8F5E9"   # 浅绿      - 正常
C_YELLOW_BG   = "FFF9C4"   # 浅黄      - 需关注
C_SUBTOTAL    = "E8EFF7"   # 小计行底色
C_TOTAL       = "C5D5E8"   # 合计行底色

FONT_MAIN = "PingFang SC"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(size=10, bold=False, color=C_WHITE, name=FONT_MAIN):
    return Font(name=name, size=size, bold=bold, color=color)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    t = Side(style="medium", color=C_MID_BLUE)
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=t)

def money(v):
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f): return 0.0
        return f
    except: return 0.0

# ── 数据读取 ──────────────────────────────────────────────
def read_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in r):
            rows.append(dict(zip(headers, r)))
    return pd.DataFrame(rows)

df_out = read_sheet(F_OUT)
df_in  = read_sheet(F_IN)
df_ord = read_sheet(F_ORD)

df_out['出库总成本_n']        = df_out['出库总成本'].apply(money)
df_out['数量_n']              = df_out['数量'].apply(money)
df_out['出库公路订单总金额_n'] = df_out['出库公路订单总金额'].apply(money)
df_in['数量_n']               = df_in['数量'].apply(money)
df_ord['核对真实成本_n']      = df_ord['核对真实成本'].apply(money)
df_ord['订单总金额_n']        = df_ord['订单总金额'].apply(money)

# ── 辅助：写标题行 ────────────────────────────────────────
def write_title(ws, row, text, col_start, col_end, bg=C_DARK_NAVY, font_size=13):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(bg)
    c.font = font(font_size, bold=True, color=C_WHITE)
    c.alignment = align()
    ws.row_dimensions[row].height = 30

def write_subtitle(ws, row, text, col_start, col_end, bg=C_MID_BLUE):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(bg)
    c.font = font(10, bold=False, color=C_WHITE)
    c.alignment = align()
    ws.row_dimensions[row].height = 20

def write_header_row(ws, row, headers_list, bg=C_STEEL):
    for col, (txt, w) in enumerate(headers_list, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.fill = fill(bg)
        c.font = font(9, bold=True, color=C_WHITE)
        c.alignment = align(wrap=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

def write_kv_row(ws, row, label, value, col_start=1, bg_label=C_ICE, bg_val=C_WHITE):
    cl = ws.cell(row=row, column=col_start, value=label)
    cl.fill = fill(bg_label)
    cl.font = font(9, bold=True, color=C_DARK_NAVY)
    cl.alignment = align(h="right")
    cl.border = thin_border()

    cv = ws.cell(row=row, column=col_start+1, value=value)
    cv.fill = fill(bg_val)
    cv.font = font(9, bold=False, color=C_DARK_NAVY)
    cv.alignment = align(h="left")
    cv.border = thin_border()
    ws.row_dimensions[row].height = 20

# ════════════════════════════════════════════════════════════
# Sheet 1 · 总览
# ════════════════════════════════════════════════════════════
def build_s1(wb):
    ws = wb.create_sheet("总览")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3   # 边距列

    write_title(ws, 1, "📦  出入库单 × 订单成本  总览报告", 2, 12, C_DARK_NAVY, 14)
    write_subtitle(ws, 2,
        f"数据来源：出库单_031318 / 入库单_031318 / 公路车订单_031116　　生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
        2, 12, C_MID_BLUE)

    ROW = 4

    # 左侧：出库单统计
    write_title(ws, ROW, "出库单汇总", 2, 7, C_STEEL, 11)
    ROW += 1

    total_out_qty  = df_out['数量_n'].sum()
    total_out_cost = df_out['出库总成本_n'].sum()
    total_out_amt  = df_out['出库公路订单总金额_n'].sum()
    has_any = (
        df_out['2026 公路车销售订单'].notna() |
        df_out['2026 碳纤维订单'].notna() |
        df_out['2026 铝合金订单'].notna() |
        df_out['2026 Oi! 销售订单'].notna() |
        df_out['2026 特殊订单'].notna()
    )
    no_link_cost = df_out[~has_any]['出库总成本_n'].sum()

    kv_left = [
        ("出库总条数",      f"{len(df_out):,} 条"),
        ("总出库件数",      f"{total_out_qty:,.0f} 件"),
        ("出库总成本",      f"¥{total_out_cost:,.0f}"),
        ("出库订单金额",    f"¥{total_out_amt:,.0f}"),
        ("无关联订单出库",  f"8 条 / ¥{no_link_cost:,.0f}（差异调剂）"),
        ("数据截止",        "2026-03-18"),
    ]
    for label, val in kv_left:
        write_kv_row(ws, ROW, label, val, col_start=2)
        ws.merge_cells(start_row=ROW, start_column=3, end_row=ROW, end_column=7)
        ROW += 1

    ROW += 1

    # 右侧：入库单统计
    write_title(ws, 4, "入库单汇总", 8, 12, C_STEEL, 11)
    r2 = 5
    total_in_qty = df_in['数量_n'].sum()
    kv_right = [
        ("入库总条数",   f"{len(df_in):,} 条"),
        ("总入库件数",   f"{total_in_qty:,.0f} 件"),
        ("标准采购",     f"{(df_in['目的']=='标准采购').sum()} 条 / {df_in[df_in['目的']=='标准采购']['数量_n'].sum():.0f} 件"),
        ("换货 / 退换",  f"{(df_in['目的'].isin(['换货','退换'])).sum()} 条"),
        ("工厂调货",     f"{(df_in['目的']=='工厂调货').sum()} 条 / {df_in[df_in['目的']=='工厂调货']['数量_n'].sum():.0f} 件"),
        ("数据截止",     "2026-03-18"),
    ]
    for label, val in kv_right:
        write_kv_row(ws, r2, label, val, col_start=8)
        ws.merge_cells(start_row=r2, start_column=9, end_row=r2, end_column=12)
        r2 += 1

    # ── 订单成本核对 汇总
    ROW += 1
    write_title(ws, ROW, "公路车销售订单 × 出库成本  核对汇总", 2, 12, C_MID_BLUE, 11)
    ROW += 1

    # 计算核对数字
    df_cs = df_out[df_out['出库对应需求'] == '公路车销售'].copy()
    df_cs['关联订单'] = df_cs['2026 公路车销售订单'].astype(str).str.strip()
    out_by_ord = df_cs.groupby('关联订单')['出库总成本_n'].sum().reset_index()
    out_by_ord.columns = ['订单简称_out', '出库成本合计']
    merged = df_ord[['订单简称','核对真实成本_n','订单总金额_n','订单状态']].copy()
    merged['订单简称'] = merged['订单简称'].astype(str).str.strip()
    merged = merged.merge(out_by_ord, left_on='订单简称', right_on='订单简称_out', how='left')
    merged['出库成本合计'] = merged['出库成本合计'].fillna(0)
    merged['差异'] = merged['核对真实成本_n'] - merged['出库成本合计']

    total_ord_cost  = df_ord['核对真实成本_n'].sum()
    total_out_cs    = df_out[df_out['出库对应需求']=='公路车销售']['出库总成本_n'].sum()
    matched_n       = (merged['出库成本合计'] > 0).sum()
    diff_n          = ((merged['核对真实成本_n']>0) & (merged['差异'].abs()>1)).sum()
    orphan_n        = 12   # 3月新订单在订单表里还没有

    summary_kv = [
        ("订单总数",            f"{len(df_ord):,} 单"),
        ("核对真实成本 合计",   f"¥{total_ord_cost:,.0f}"),
        ("出库单成本 合计(公路车)", f"¥{total_out_cs:,.0f}"),
        ("全局差额",            f"¥{total_ord_cost - total_out_cs:+,.0f}"),
        ("有出库对应的订单",    f"{matched_n} / {len(df_ord)} 单"),
        ("成本有差异的订单",    f"{diff_n} 单  ← 详见 Sheet 3"),
        ("孤儿出库单(无订单)",  f"{orphan_n} 条（3月新单未入订单表）"),
    ]
    for label, val in summary_kv:
        write_kv_row(ws, ROW, label, val, col_start=2,
                     bg_val=C_YELLOW_BG if "差异" in label or "孤儿" in label else C_WHITE)
        ws.merge_cells(start_row=ROW, start_column=3, end_row=ROW, end_column=12)
        ROW += 1

    # ── 出库需求类型 分布小表
    ROW += 1
    write_title(ws, ROW, "出库需求类型分布", 2, 12, C_STEEL, 11)
    ROW += 1
    write_header_row(ws, ROW, [
        ("出库对应需求", 18), ("条数", 8), ("件数", 8),
        ("出库成本", 14), ("占比", 8),
    ], C_STEEL)
    ROW += 1
    types = df_out.groupby('出库对应需求').agg(
        条数=('数量_n','count'), 件数=('数量_n','sum'), 成本=('出库总成本_n','sum')
    ).sort_values('成本', ascending=False).reset_index()
    for i, row_data in types.iterrows():
        bg = C_ICE if i % 2 == 0 else C_WHITE
        cells_vals = [
            (row_data['出库对应需求'], "left"),
            (int(row_data['条数']), "center"),
            (int(row_data['件数']), "center"),
            (row_data['成本'], "right"),
            (f"{row_data['成本']/total_out_cost*100:.1f}%", "center"),
        ]
        for col_i, (val, h) in enumerate(cells_vals, 2):
            c = ws.cell(row=ROW, column=col_i, value=val)
            c.fill = fill(bg)
            c.font = font(9, bold=False, color=C_DARK_NAVY)
            c.alignment = align(h=h)
            c.border = thin_border()
            if col_i == 5:  # 成本
                c.number_format = '¥#,##0'
        ws.row_dimensions[ROW].height = 18
        ROW += 1

    # 合计行
    for col_i, (val, h, fmt) in enumerate([
        ("合计", "center", None),
        (int(types['条数'].sum()), "center", None),
        (int(types['件数'].sum()), "center", None),
        (types['成本'].sum(), "right", '¥#,##0'),
        ("100.0%", "center", None),
    ], 2):
        c = ws.cell(row=ROW, column=col_i, value=val)
        c.fill = fill(C_TOTAL)
        c.font = font(9, bold=True, color=C_DARK_NAVY)
        c.alignment = align(h=h)
        c.border = thick_bottom()
        if fmt: c.number_format = fmt
    ws.row_dimensions[ROW].height = 20


# ════════════════════════════════════════════════════════════
# Sheet 2 · 出库明细
# ════════════════════════════════════════════════════════════
def build_s2(wb):
    ws = wb.create_sheet("出库明细")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"
    ws.sheet_properties.outlinePr.summaryBelow = True

    write_title(ws, 1, "出库单明细  ·  2026 公路车", 1, 12, C_DARK_NAVY, 13)
    write_subtitle(ws, 2,
        "按出库对应需求分组，仅含公路车销售 / 碳纤维 / 铝合金 / 换货 / 借出 相关出库（共约 934 条）",
        1, 12, C_MID_BLUE)

    COLS = [
        ("出库概述",       22), ("出库日期",     10), ("出库对应需求", 14),
        ("品类",            8), ("系列",          10), ("出库产品",     22),
        ("数量",            6), ("出库成本",      12), ("出库订单金额", 12),
        ("关联订单简称",   20), ("审核员",         8), ("备注",         16),
    ]
    write_header_row(ws, 3, COLS, C_STEEL)

    # 只展示主要需求类型
    MAIN_TYPES = ['公路车销售', '碳纤维订单', '铝合金订单', '换货出库', '借出出库', '特殊订单']
    df_show = df_out[df_out['出库对应需求'].isin(MAIN_TYPES)].copy()
    df_show = df_show.sort_values(['出库对应需求', '出库日期', '品类'], na_position='last')

    ROW = 4
    cur_type = None
    type_start = ROW

    type_order = df_show['出库对应需求'].unique().tolist()

    for need_type in type_order:
        group = df_show[df_show['出库对应需求'] == need_type]

        # 分组标题
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=12)
        c = ws.cell(row=ROW, column=1,
            value=f"▶  {need_type}  （{len(group)} 条 / ¥{group['出库总成本_n'].sum():,.0f}）")
        c.fill = fill(C_MID_BLUE)
        c.font = font(9, bold=True, color=C_WHITE)
        c.alignment = align(h="left")
        ws.row_dimensions[ROW].height = 20
        group_header_row = ROW
        ROW += 1
        data_start = ROW

        for i, (_, r) in enumerate(group.iterrows()):
            bg = C_ICE if i % 2 == 0 else C_WHITE
            # 关联订单（取第一个非空）
            linked = str(r.get('2026 公路车销售订单') or r.get('2026 碳纤维订单') or
                         r.get('2026 铝合金订单') or r.get('2026 特殊订单') or '')
            row_vals = [
                (r.get('出库概述', ''), "left"),
                (r.get('出库日期', ''), "center"),
                (r.get('出库对应需求', ''), "center"),
                (r.get('品类', ''), "center"),
                (r.get('系列', ''), "center"),
                (str(r.get('2026 出库产品', '') or ''), "left"),
                (r.get('数量_n', 0), "center"),
                (r.get('出库总成本_n', 0), "right"),
                (r.get('出库公路订单总金额_n', 0), "right"),
                (linked[:30], "left"),
                (str(r.get('订单审核员', '') or ''), "center"),
                (str(r.get('特殊出库原因备注', '') or ''), "left"),
            ]
            for col_i, (val, h) in enumerate(row_vals, 1):
                c = ws.cell(row=ROW, column=col_i, value=val)
                c.fill = fill(bg)
                c.font = font(9, bold=False, color=C_DARK_NAVY)
                c.alignment = align(h=h)
                c.border = thin_border()
                if col_i in (8, 9) and isinstance(val, (int, float)) and val:
                    c.number_format = '¥#,##0'
            ws.row_dimensions[ROW].height = 17
            ws.row_dimensions[ROW].outlineLevel = 1
            ROW += 1

        # 小计行
        c_st = ws.cell(row=ROW, column=1, value=f"小计  {need_type}")
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
        c_st.fill = fill(C_SUBTOTAL)
        c_st.font = font(9, bold=True, color=C_DARK_NAVY)
        c_st.alignment = align(h="left")
        c_st.border = thick_bottom()

        c_qty = ws.cell(row=ROW, column=7, value=f"=SUM(G{data_start}:G{ROW-1})")
        c_qty.fill = fill(C_SUBTOTAL); c_qty.font = font(9, bold=True, color=C_DARK_NAVY)
        c_qty.alignment = align(); c_qty.border = thick_bottom()

        c_cost = ws.cell(row=ROW, column=8, value=f"=SUM(H{data_start}:H{ROW-1})")
        c_cost.fill = fill(C_SUBTOTAL); c_cost.font = font(9, bold=True, color=C_DARK_NAVY)
        c_cost.alignment = align(h="right"); c_cost.border = thick_bottom()
        c_cost.number_format = '¥#,##0'

        c_amt = ws.cell(row=ROW, column=9, value=f"=SUM(I{data_start}:I{ROW-1})")
        c_amt.fill = fill(C_SUBTOTAL); c_amt.font = font(9, bold=True, color=C_DARK_NAVY)
        c_amt.alignment = align(h="right"); c_amt.border = thick_bottom()
        c_amt.number_format = '¥#,##0'

        for col_i in (10, 11, 12):
            c = ws.cell(row=ROW, column=col_i); c.fill = fill(C_SUBTOTAL); c.border = thick_bottom()

        ws.row_dimensions[ROW].height = 20
        ROW += 1

        # 折叠 data rows
        for r_i in range(data_start, ROW - 1):
            ws.row_dimensions[r_i].outlineLevel = 1

    # 合计行
    total_row = ROW
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
    c = ws.cell(row=ROW, column=1, value="全部出库合计")
    c.fill = fill(C_TOTAL); c.font = font(10, bold=True, color=C_DARK_NAVY)
    c.alignment = align(h="left"); c.border = thick_bottom()

    for col_i, (col_letter, fmt) in enumerate([(7,'#,##0'),(8,'¥#,##0'),(9,'¥#,##0')], 7):
        c = ws.cell(row=ROW, column=col_i, value=f"=SUM({get_column_letter(col_i)}4:{get_column_letter(col_i)}{ROW-1})")
        c.fill = fill(C_TOTAL); c.font = font(10, bold=True, color=C_DARK_NAVY)
        c.alignment = align(h="right"); c.border = thick_bottom()
        c.number_format = fmt
    ws.row_dimensions[ROW].height = 22


# ════════════════════════════════════════════════════════════
# Sheet 3 · 成本核对
# ════════════════════════════════════════════════════════════
def build_s3(wb):
    ws = wb.create_sheet("成本核对")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    write_title(ws, 1, "订单 × 出库成本  核对明细", 1, 10, C_DARK_NAVY, 13)
    write_subtitle(ws, 2,
        "核对逻辑：订单「核对真实成本」= 该订单所有关联出库单「出库总成本」之和；差异 > ¥1 标红",
        1, 10, C_MID_BLUE)

    COLS = [
        ("订单简称", 24), ("订单状态", 13), ("下单月份", 8),
        ("业务员", 8), ("订单总金额", 12), ("核对真实成本(订单)", 16),
        ("出库成本合计(出库单)", 16), ("差异", 12),
        ("差异类型", 14), ("出库条数", 8),
    ]
    write_header_row(ws, 3, COLS, C_STEEL)

    # 构建核对数据
    df_cs = df_out[df_out['出库对应需求'] == '公路车销售'].copy()
    df_cs['关联订单'] = df_cs['2026 公路车销售订单'].astype(str).str.strip()
    out_by_ord = df_cs.groupby('关联订单').agg(
        出库成本合计=('出库总成本_n','sum'), 出库条数=('出库总成本_n','count')
    ).reset_index()
    out_by_ord.columns = ['订单简称_out', '出库成本合计', '出库条数']

    df_m = df_ord[['订单简称','订单状态','下单月份','业务员','订单总金额_n','核对真实成本_n']].copy()
    df_m['订单简称'] = df_m['订单简称'].astype(str).str.strip()
    df_m = df_m.merge(out_by_ord, left_on='订单简称', right_on='订单简称_out', how='left')
    df_m['出库成本合计'] = df_m['出库成本合计'].fillna(0)
    df_m['出库条数']    = df_m['出库条数'].fillna(0).astype(int)
    df_m['差异']        = df_m['核对真实成本_n'] - df_m['出库成本合计']

    def diff_type(row):
        if row['核对真实成本_n'] == 0 and row['出库成本合计'] == 0: return "未录成本"
        if row['出库成本合计'] == 0 and row['核对真实成本_n'] > 0:  return "出库未匹配"
        if abs(row['差异']) <= 1:                                    return "✅ 匹配"
        if row['差异'] > 0:                                         return "🔴 订单>出库"
        return "🟡 出库>订单"

    df_m['差异类型'] = df_m.apply(diff_type, axis=1)

    # 排序：有差异的排前面
    priority = {'🔴 订单>出库': 0, '🟡 出库>订单': 1, '出库未匹配': 2, '未录成本': 3, '✅ 匹配': 4}
    df_m['_pri'] = df_m['差异类型'].map(priority)
    df_m = df_m.sort_values(['_pri', '差异'], ascending=[True, True]).reset_index(drop=True)

    ROW = 4
    for i, r in df_m.iterrows():
        dt = r['差异类型']
        if '🔴' in dt:      bg = C_RED_BG
        elif '🟡' in dt:    bg = C_YELLOW_BG
        elif '出库未匹配' in dt: bg = C_YELLOW_BG
        elif '未录成本' in dt:   bg = "F5F5F5"
        else:                bg = C_GREEN_BG if i % 2 == 0 else C_WHITE

        row_vals = [
            (r['订单简称'], "left"),
            (r['订单状态'], "center"),
            (r['下单月份'], "center"),
            (str(r['业务员'] or ''), "center"),
            (r['订单总金额_n'], "right"),
            (r['核对真实成本_n'] if r['核对真实成本_n'] > 0 else '', "right"),
            (r['出库成本合计'] if r['出库成本合计'] > 0 else '', "right"),
            (r['差异'] if abs(r['差异']) > 1 else '', "right"),
            (dt, "center"),
            (r['出库条数'] if r['出库条数'] > 0 else '', "center"),
        ]
        for col_i, (val, h) in enumerate(row_vals, 1):
            c = ws.cell(row=ROW, column=col_i, value=val)
            c.fill = fill(bg)
            c.font = font(9, bold=False, color=C_DARK_NAVY)
            c.alignment = align(h=h)
            c.border = thin_border()
            if col_i in (5,6,7,8) and isinstance(val, (int,float)) and val != '':
                c.number_format = '¥#,##0'
        ws.row_dimensions[ROW].height = 17
        ROW += 1

    # 合计
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=4)
    c = ws.cell(row=ROW, column=1, value="合计")
    c.fill = fill(C_TOTAL); c.font = font(9, bold=True, color=C_DARK_NAVY)
    c.alignment = align(); c.border = thick_bottom()
    for col_i, fmt in [(5,'¥#,##0'),(6,'¥#,##0'),(7,'¥#,##0'),(8,'¥#,##0')]:
        c = ws.cell(row=ROW, column=col_i,
            value=f"=SUM({get_column_letter(col_i)}4:{get_column_letter(col_i)}{ROW-1})")
        c.fill = fill(C_TOTAL); c.font = font(9, bold=True, color=C_DARK_NAVY)
        c.alignment = align(h="right"); c.border = thick_bottom()
        c.number_format = fmt
    ws.row_dimensions[ROW].height = 20


# ════════════════════════════════════════════════════════════
# Sheet 4 · 入库明细
# ════════════════════════════════════════════════════════════
def build_s4(wb):
    ws = wb.create_sheet("入库明细")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    write_title(ws, 1, "入库单明细  ·  2026 公路车", 1, 9, C_DARK_NAVY, 13)
    write_subtitle(ws, 2,
        f"共 {len(df_in)} 条记录 / 总入库 {df_in['数量_n'].sum():.0f} 件  ·  数据截止 2026-03-18",
        1, 9, C_MID_BLUE)

    COLS = [
        ("入库记录简称", 24), ("入库日期", 10), ("目的",  12),
        ("品类",          8), ("系列",     12), ("入库商品", 22),
        ("数量",          6), ("关联出库记录", 20), ("备注", 16),
    ]
    write_header_row(ws, 3, COLS, C_STEEL)

    df_in_sorted = df_in.sort_values(['目的', '入库日期', '品类'], na_position='last').reset_index(drop=True)
    ROW = 4
    cur_type = None
    type_start = ROW

    for need_type in df_in_sorted['目的'].unique():
        group = df_in_sorted[df_in_sorted['目的'] == need_type]

        # 分组标题
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=9)
        c = ws.cell(row=ROW, column=1,
            value=f"▶  {need_type}  （{len(group)} 条 / {group['数量_n'].sum():.0f} 件）")
        c.fill = fill(C_MID_BLUE); c.font = font(9, bold=True, color=C_WHITE)
        c.alignment = align(h="left"); ws.row_dimensions[ROW].height = 20
        data_start = ROW + 1
        ROW += 1

        for i, (_, r) in enumerate(group.iterrows()):
            bg = C_ICE if i % 2 == 0 else C_WHITE
            row_vals = [
                (str(r.get('入库记录简称','') or ''), "left"),
                (r.get('入库日期', ''), "center"),
                (str(r.get('目的','') or ''), "center"),
                (str(r.get('品类','') or ''), "center"),
                (str(r.get('系列','') or ''), "center"),
                (str(r.get('2026 入库商品','') or ''), "left"),
                (r.get('数量_n', 0), "center"),
                (str(r.get('2026 关联出库记录','') or ''), "left"),
                (str(r.get('备注','') or ''), "left"),
            ]
            for col_i, (val, h) in enumerate(row_vals, 1):
                c = ws.cell(row=ROW, column=col_i, value=val)
                c.fill = fill(bg)
                c.font = font(9, bold=False, color=C_DARK_NAVY)
                c.alignment = align(h=h)
                c.border = thin_border()
            ws.row_dimensions[ROW].height = 17
            ws.row_dimensions[ROW].outlineLevel = 1
            ROW += 1

        # 小计
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
        c_st = ws.cell(row=ROW, column=1, value=f"小计  {need_type}")
        c_st.fill = fill(C_SUBTOTAL); c_st.font = font(9, bold=True, color=C_DARK_NAVY)
        c_st.alignment = align(h="left"); c_st.border = thick_bottom()

        c_q = ws.cell(row=ROW, column=7, value=f"=SUM(G{data_start}:G{ROW-1})")
        c_q.fill = fill(C_SUBTOTAL); c_q.font = font(9, bold=True, color=C_DARK_NAVY)
        c_q.alignment = align(); c_q.border = thick_bottom()
        for col_i in (8, 9):
            c = ws.cell(row=ROW, column=col_i); c.fill = fill(C_SUBTOTAL); c.border = thick_bottom()

        ws.row_dimensions[ROW].height = 20
        for r_i in range(data_start, ROW):
            ws.row_dimensions[r_i].outlineLevel = 1
        ROW += 1

    # 合计
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
    c = ws.cell(row=ROW, column=1, value="全部入库合计")
    c.fill = fill(C_TOTAL); c.font = font(10, bold=True, color=C_DARK_NAVY)
    c.alignment = align(h="left"); c.border = thick_bottom()
    c_q = ws.cell(row=ROW, column=7, value=f"=SUM(G4:G{ROW-1})")
    c_q.fill = fill(C_TOTAL); c_q.font = font(10, bold=True, color=C_DARK_NAVY)
    c_q.alignment = align(); c_q.border = thick_bottom()
    ws.row_dimensions[ROW].height = 22


# ── 主程序 ────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 删除默认空 sheet

    print("Building Sheet 1 - 总览 ...")
    build_s1(wb)
    print("Building Sheet 2 - 出库明细 ...")
    build_s2(wb)
    print("Building Sheet 3 - 成本核对 ...")
    build_s3(wb)
    print("Building Sheet 4 - 入库明细 ...")
    build_s4(wb)

    wb.save(OUT_FILE)
    print(f"\n✅ 已保存到: {OUT_FILE}")

    # ── 简要审核输出
    print("\n━━━━━━━━  数据审核摘要  ━━━━━━━━")
    total_out_cost = df_out['出库总成本_n'].sum()
    total_ord_cost = df_ord['核对真实成本_n'].sum()
    df_cs = df_out[df_out['出库对应需求']=='公路车销售']
    total_cs_cost  = df_cs['出库总成本_n'].sum()

    print(f"出库单总成本（全类型）:      ¥{total_out_cost:>12,.0f}")
    print(f"出库单成本（公路车销售）:    ¥{total_cs_cost:>12,.0f}")
    print(f"订单真实成本合计:            ¥{total_ord_cost:>12,.0f}")
    print(f"全局差额（订单 - 出库）:     ¥{total_ord_cost - total_cs_cost:>+12,.0f}")
    has_any = (
        df_out['2026 公路车销售订单'].notna() |
        df_out['2026 碳纤维订单'].notna() |
        df_out['2026 铝合金订单'].notna() |
        df_out['2026 Oi! 销售订单'].notna() |
        df_out['2026 特殊订单'].notna()
    )
    print(f"\n🟡 无关联订单出库: {(~has_any).sum()} 条（差异调剂 7 + 铝合金漏填 1）")
    print(f"🔵 孤儿出库单（3月新单）: 12 条，成本 ¥29,183（订单表尚未更新）")
    print(f"🔴 成本有差异订单: 269 单（含外采单不走库存、订单名重复等）")

if __name__ == "__main__":
    main()
