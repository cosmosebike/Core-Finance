#!/usr/bin/env python3
"""
公路车库存_20260313_v3.xlsx — 构建脚本
STYLE1 权威咨询风 · 全公式驱动 · 已修正所有 v2 问题
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os, re, math

# ═══════════════════════════════════════════════════════════════
# 时间戳解析：从文件名 _MMDDHHH 中提取日期时间
# ═══════════════════════════════════════════════════════════════
def parse_ts(filepath):
    """从路径中提取 MMDDHHH 时间戳，只返回日期显示字符串（不含时间）
    例如 031316 → ('3/13', '3/13')"""
    m = re.search(r'_(\d{6})\.xlsx', os.path.basename(filepath))
    if not m:
        return ('—', '—')
    ts = m.group(1)   # e.g. '031316'
    mm, dd = ts[0:2], ts[2:4]
    date_str = f"{int(mm)}/{int(dd)}"
    return (date_str, date_str)

# ═══════════════════════════════════════════════════════════════
# STYLE1 色彩常量
# ═══════════════════════════════════════════════════════════════
C_PAPER      = "FAF9F6"
C_WHITE      = "FFFFFF"
C_NEAR_BLACK = "111111"
C_MUTED      = "555555"
C_LIGHT_GREY = "CCCCCC"
C_ROW_STRIPE = "F5F5F5"
C_TOTAL_BG   = "E8E8E8"
C_BURGUNDY   = "8B1A1A"
C_AMBER      = "FFF8E8"   # 新增SKU / (待确认)背景

# ═══════════════════════════════════════════════════════════════
# 文件路径
# ═══════════════════════════════════════════════════════════════
BASE      = "/Users/miaworkbook/Desktop/核心数据库"
SNAP      = f"{BASE}/数据快照"
FILE_CURR = f"{SNAP}/宇宙销售统计中心_2026 商品SKU_031316.xlsx"
FILE_PREV = f"{SNAP}/宇宙销售统计中心_2026 商品SKU_031214.xlsx"
FILE_V6   = f"{BASE}/库存报表/公路车库存_20260312_v6.xlsx"
FILE_QICHU = "/Users/miaworkbook/Downloads/2026清货KPI管理表_表1表2.xlsx"
OUT       = f"{BASE}/库存报表/公路车库存_20260313_v14.xlsx"


# ═══════════════════════════════════════════════════════════════
# 样式辅助函数
# ═══════════════════════════════════════════════════════════════
def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mk_font(bold=False, size=10, color=C_NEAR_BLACK, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mk_align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def mk_border_thin():
    s = Side(style="thin", color=C_LIGHT_GREY)
    return Border(top=s, bottom=s)

def mk_border_medium_top():
    return Border(top=Side(style="medium", color=C_NEAR_BLACK))

def mk_border_medium_bottom():
    return Border(bottom=Side(style="medium", color=C_NEAR_BLACK))

# ═══════════════════════════════════════════════════════════════
# 数字格式
# ═══════════════════════════════════════════════════════════════
FMT_INT   = '#,##0'
FMT_COST  = '#,##0'
FMT_PCT   = '0.0%'

# ═══════════════════════════════════════════════════════════════
# 排序规则
# ═══════════════════════════════════════════════════════════════
LEVEL_ORDER = [
    'Dogma 系列', 'Pinarello 清理', '清理库存',
    '正常商品', 'DA/Zipp 轮组', '换货回商品',
    '自营产品', '套件类', 'FSA 系列',
    '配件', '铺货商品',
]
LEVEL_ICONS = {
    'Dogma 系列':    '🔵',
    'Pinarello 清理':'🔴',
    '清理库存':      '🔴',
    '正常商品':      '🟢',
    'DA/Zipp 轮组':  '🔵',
    '换货回商品':    '🟡',
    'FSA 系列':      '🟡',
    '自营产品':      '🟢',
    '套件类':        '🟢',
    '配件':          '⚪',
    '铺货商品':      '⚪',
}
# 品类优先级（整车/车架先，轮组，外胎，其余）
CATEGORY_PRIORITY = {
    # 整车/车架/轮组 类
    '整车': 1, '车架': 2,
    '轮组': 3,
    '外胎': 4,
    # 传动系统
    '盘片': 5, '碟片组': 5,
    '曲柄组': 6, '牙盘组': 6,
    '功率计': 7,
    # 坐骑/操控
    '座垫': 8,
    '车把一体把': 9,
    # 套件
    '套件组': 10, '套件散装': 10, '套件链条': 10,
    '链条': 10,
    # 小零件
    '后拨导轮': 11,
    '内胎': 12, '把带': 13, '中轴': 14,
    '锁踏': 15, '脚踏': 15,
    # 配件类（头盔/眼镜/骑行服等）
    '头盔': 20,
    '眼镜': 21,
    '骑行服': 22,
    '其他小配件': 23,
    # 骑行台
    '骑行台': 30,
}
# Sheet3 高流通级别（排除清理类）
HIGH_CIRC_LEVELS = {
    '正常商品', 'DA/Zipp 轮组', '换货回商品', 'FSA 系列',
    '自营产品', '套件类', '配件', '铺货商品',
}

def level_idx(lvl):
    try: return LEVEL_ORDER.index(str(lvl).strip() if lvl else '')
    except ValueError: return 99

def cat_idx(cat):
    return CATEGORY_PRIORITY.get(str(cat).strip() if cat else '', 50)

def to_num(v):
    import math
    try:
        r = float(v) if v is not None else 0.0
        return 0.0 if math.isnan(r) or math.isinf(r) else r
    except:
        return 0.0

# ═══════════════════════════════════════════════════════════════
# 数据加载
# ═══════════════════════════════════════════════════════════════
def load_df(filepath, sheet_idx=0):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sname = wb.sheetnames[sheet_idx]
    ws = wb[sname]
    headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row[:5]):
            data.append(dict(zip(headers, row)))
    return pd.DataFrame(data)

def load_qichu_data():
    """从 2026清货KPI管理表 读取期初(2/8)件数和成本"""
    wb = openpyxl.load_workbook(FILE_QICHU, data_only=True)
    ws = wb['表1-库存分类总览']
    headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    qichu = {}  # normalized_sku -> {'qty': n, 'cost': n, 'total_cost': n}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        sku = str(d.get('SKU名称', '') or '').strip()
        if not sku:
            continue
        qty = to_num(d.get('库存件数', 0))
        cost = to_num(d.get('实际成本', 0))
        total = to_num(d.get('总成本', 0))
        if total == 0:
            total = qty * cost
        qichu[normalize_sku(sku)] = {'qty': qty, 'cost': cost, 'total_cost': total}
    return qichu


def load_v6_person_map():
    """从 v6 Sheet4 读取 SKU -> 负责人 映射"""
    wb = openpyxl.load_workbook(FILE_V6, data_only=True)
    ws = wb['库存清理跟踪']
    person = None
    mapping = {}  # normalized_sku -> person
    for row in ws.iter_rows(min_row=1, values_only=True):
        b = row[1] if len(row) > 1 else None
        if b is None:
            continue
        b_str = str(b).strip()
        # 人员标题行 (▸ 张振 — ...)
        if b_str.startswith('▸'):
            parts = b_str.split('—')[0].replace('▸', '').strip()
            person = parts.strip()
        elif person and len(b_str) > 4 and '使用方式' not in b_str and '数据基准' not in b_str and '注：' not in b_str and '数据来源' not in b_str and '⚠' not in b_str and '✅' not in b_str:
            mapping[normalize_sku(b_str)] = person
    return mapping

def normalize_sku(s):
    """标准化SKU名称用于匹配"""
    return ' '.join(str(s).strip().split()).lower()

# ═══════════════════════════════════════════════════════════════
# 通用写入辅助
# ═══════════════════════════════════════════════════════════════
def set_no_gridlines(ws):
    ws.sheet_view.showGridLines = False

def write_title_block(ws, row, title, subtitle, col_end=12):
    """写标题(row) + 副标题(row+1) + 空白行(row+2)"""
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=2, value=title)
    c.fill = mk_fill(C_NEAR_BLACK); c.font = mk_font(bold=True, size=14, color=C_WHITE)
    c.alignment = mk_align("left"); ws.row_dimensions[row].height = 32

    ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=col_end)
    c2 = ws.cell(row=row+1, column=2, value=subtitle)
    c2.fill = mk_fill(C_PAPER); c2.font = mk_font(size=9, color=C_MUTED)
    c2.alignment = mk_align("left"); ws.row_dimensions[row+1].height = 18

    for col in range(2, col_end+1):
        ws.cell(row=row+2, column=col).fill = mk_fill(C_WHITE)
    ws.row_dimensions[row+2].height = 6

def write_col_headers(ws, row, headers, col_start=2):
    ws.row_dimensions[row].height = 28
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start+i, value=h)
        c.fill = mk_fill(C_NEAR_BLACK)
        c.font = mk_font(bold=True, size=10, color=C_WHITE)
        c.alignment = mk_align("center", wrap=True)
        c.border = mk_border_medium_bottom()

def write_level_header(ws, row, label, col_end=12):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=2, value=label)
    c.fill = mk_fill(C_ROW_STRIPE)
    c.font = mk_font(bold=True, size=10)
    c.alignment = mk_align("left")
    ws.row_dimensions[row].height = 22

def write_data_row(ws, row, values, col_start=2, stripe=False):
    bg = C_ROW_STRIPE if stripe else C_WHITE
    ws.row_dimensions[row].height = 22
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start+i, value=val)
        c.fill = mk_fill(bg)
        c.font = mk_font(size=10)
        c.border = mk_border_thin()

def write_subtotal_row(ws, row, values, col_start=2, col_end=12, label="小计"):
    ws.row_dimensions[row].height = 22
    for col in range(col_start, col_end+1):
        c = ws.cell(row=row, column=col)
        c.fill = mk_fill(C_TOTAL_BG)
        c.font = mk_font(bold=True, size=10)
        c.alignment = mk_align("center")
        c.border = mk_border_medium_top()
    # Write values
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start+i, value=val)

def write_note(ws, row, text, col_end=12):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=2, value=text)
    c.fill = mk_fill(C_WHITE); c.font = mk_font(size=9, color=C_MUTED, italic=True)
    c.alignment = mk_align("left"); ws.row_dimensions[row].height = 16

def set_col_widths(ws, widths):
    """widths: {col_letter: width}"""
    ws.column_dimensions['A'].width = 2
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════
# SHEET 1: 总览
# ═══════════════════════════════════════════════════════════════
def build_sheet1(wb, df, ts_curr):
    ws = wb.active
    ws.title = '总览'
    ws.sheet_properties.tabColor = C_NEAR_BLACK
    set_no_gridlines(ws)
    set_col_widths(ws, {'B': 20, 'C': 12, 'D': 12, 'E': 16, 'F': 12})

    write_title_block(ws, 2,
        f'宇宙销售统计中心  ·  公路车库存总览  ·  {ts_curr}',
        f'数据基准：{ts_curr}  ·  来源：{os.path.basename(FILE_CURR)}',
        col_end=6)

    write_col_headers(ws, 5, ['库存级别', 'SKU数量', '总件数', '总成本 (¥)', '占总成本比'], col_start=2)

    # 统计各级别
    summary = []
    total_cost_all = 0
    level_rows = []
    for lvl in LEVEL_ORDER:
        grp = df[df['库存级别'] == lvl]
        if len(grp) == 0:
            continue
        sku_cnt = len(grp)
        total_units = sum(
            (to_num(r.get('2026 滨江仓库数量', 0)) + to_num(r.get('2026 常州数量', 0)))
            for _, r in grp.iterrows()
        )
        total_cost = sum(
            (to_num(r.get('2026 滨江仓库数量', 0)) + to_num(r.get('2026 常州数量', 0))) * to_num(r.get('实际成本价', 0))
            for _, r in grp.iterrows()
        )
        total_cost_all += total_cost
        icon = LEVEL_ICONS.get(lvl, '⚪')
        summary.append((f'{icon} {lvl}', sku_cnt, total_units, total_cost))
        level_rows.append(len(summary) + 5)  # row number

    stripe = False
    data_start_row = 6
    for i, (lvl_label, sku_cnt, units, cost) in enumerate(summary):
        r = data_start_row + i
        stripe = i % 2 == 1
        bg = C_ROW_STRIPE if stripe else C_WHITE
        ws.row_dimensions[r].height = 22
        for col in range(2, 7):
            c = ws.cell(row=r, column=col)
            c.fill = mk_fill(bg)
            c.border = mk_border_thin()

        ws.cell(row=r, column=2, value=lvl_label).font = mk_font(bold=False, size=10)
        ws.cell(row=r, column=2).alignment = mk_align("left")
        ws.cell(row=r, column=3, value=sku_cnt).number_format = FMT_INT
        ws.cell(row=r, column=4, value=units).number_format = FMT_INT
        ws.cell(row=r, column=5, value=cost).number_format = FMT_COST
        # 占比 formula (后面补)
        ws.cell(row=r, column=6, value=None)

    total_row = data_start_row + len(summary)
    write_subtotal_row(ws, total_row,
        ['合计',
         f'=SUM(C{data_start_row}:C{total_row-1})',
         f'=SUM(D{data_start_row}:D{total_row-1})',
         f'=SUM(E{data_start_row}:E{total_row-1})',
         None],
        col_start=2, col_end=6)
    ws.cell(row=total_row, column=3).number_format = FMT_INT
    ws.cell(row=total_row, column=4).number_format = FMT_INT
    ws.cell(row=total_row, column=5).number_format = FMT_COST

    # 写占比 (=各行成本/合计成本)
    for i in range(len(summary)):
        r = data_start_row + i
        c = ws.cell(row=r, column=6, value=f'=E{r}/E{total_row}')
        c.number_format = FMT_PCT

    note_row = total_row + 2
    write_note(ws, note_row,
        f'数据来源：{os.path.basename(FILE_CURR)}  |  统计时间：{ts_curr}',
        col_end=6)


# ═══════════════════════════════════════════════════════════════
# SHEET 2: 库存SKU明细
# ═══════════════════════════════════════════════════════════════
S2_COLS = ['品类', 'SKU名称', '系列', '库存级别', '滨江', '常州', '总件数', '成本价', '总成本', '清理底价', '处理方式']

def build_sheet2(wb, df, ts_curr):
    ws = wb.create_sheet('库存SKU明细')
    ws.sheet_properties.tabColor = "2C3E6B"
    ws.sheet_properties.outlinePr.summaryBelow = True   # 小计在组下方，+ 号在小计行一侧
    set_no_gridlines(ws)
    set_col_widths(ws, {
        'B': 12, 'C': 32, 'D': 14, 'E': 14,
        'F': 8, 'G': 8, 'H': 8, 'I': 11, 'J': 13,
        'K': 11, 'L': 14,
    })

    write_title_block(ws, 2,
        f'库存 SKU 明细  ·  {ts_curr}',
        f'来源：{os.path.basename(FILE_CURR)} · 按级别→品类→系列→名称排序',
        col_end=12)
    write_col_headers(ws, 5, S2_COLS, col_start=2)

    # 排序：库存级别 → 品类 → 系列 → SKU名称
    df2 = df.copy()
    df2['_lvl']    = df2['库存级别'].apply(level_idx)
    df2['_cat']    = df2['品类'].apply(cat_idx)
    df2['_series'] = df2['系列'].apply(lambda x: str(x or '').strip().lower())
    df2['_name']   = df2['名称'].apply(lambda x: str(x or '').strip().lower())
    df2 = df2.sort_values(
        ['_lvl', '_cat', '_series', '_name'],
        ascending=[True, True, True, True]
    )

    row = 6
    prev_level = None
    level_start = {}   # level -> first data row
    level_end = {}     # level -> last data row

    for _, r in df2.iterrows():
        lvl = str(r.get('库存级别', '') or '').strip()

        if lvl != prev_level:
            if prev_level is not None:
                level_end[prev_level] = row - 1
                # 写小计行
                ls = level_start[prev_level]
                le = level_end[prev_level]
                _write_s2_subtotal(ws, row, prev_level, ls, le)
                row += 1

            icon = LEVEL_ICONS.get(lvl, '⚪')
            write_level_header(ws, row, f'{icon} {lvl}', col_end=12)
            row += 1
            level_start[lvl] = row
            prev_level = lvl

        # 总件数和总成本用公式
        binjiang = to_num(r.get('2026 滨江仓库数量', 0))
        changzhou = to_num(r.get('2026 常州数量', 0))
        cost_price = to_num(r.get('实际成本价', 0))

        h_col = get_column_letter(2 + 6)   # 'H' = 总件数
        i_col = get_column_letter(2 + 7)   # 'I' = 成本价
        j_col = get_column_letter(2 + 8)   # 'J' = 总成本

        stripe = (row % 2 == 0)
        bg = C_ROW_STRIPE if stripe else C_WHITE
        ws.row_dimensions[row].height = 22
        ws.row_dimensions[row].outlineLevel = 1   # 数据行归入组内，可折叠

        vals = [
            str(r.get('品类', '') or '').strip(),
            str(r.get('名称', '') or '').strip(),
            str(r.get('系列', '') or '').strip(),
            str(r.get('库存级别', '') or '').strip(),
            binjiang,
            changzhou,
            f'=F{row}+G{row}',       # 总件数公式
            cost_price,
            f'=H{row}*I{row}',       # 总成本公式
            r.get('清理底价', None),
            str(r.get('处理方式', '') or '').strip(),
        ]

        for i, val in enumerate(vals):
            col = 2 + i
            c = ws.cell(row=row, column=col, value=val)
            c.fill = mk_fill(bg)
            c.font = mk_font(size=10)
            c.border = mk_border_thin()
            # 对齐：SKU名称左对齐，其余居中
            if i == 1 or i == 10:
                c.alignment = mk_align("left")
            else:
                c.alignment = mk_align("center")

        # 数字格式
        ws.cell(row=row, column=2+4).number_format = FMT_INT  # 滨江
        ws.cell(row=row, column=2+5).number_format = FMT_INT  # 常州
        ws.cell(row=row, column=2+6).number_format = FMT_INT  # 总件数
        ws.cell(row=row, column=2+7).number_format = FMT_COST # 成本价
        ws.cell(row=row, column=2+8).number_format = FMT_COST # 总成本
        ws.cell(row=row, column=2+9).number_format = FMT_COST # 清理底价

        row += 1

    # 最后一个级别的小计
    if prev_level:
        level_end[prev_level] = row - 1
        ls = level_start[prev_level]
        le = level_end[prev_level]
        _write_s2_subtotal(ws, row, prev_level, ls, le)
        row += 1

    # 总合计行
    all_data_rows = [(v, level_end[k]) for k, v in level_start.items()]
    # 用 SUM of J column across entire data range
    ws.row_dimensions[row].height = 22
    write_subtotal_row(ws, row,
        ['总计', None, None, None, None, None,
         f'=SUMIF(E6:E{row-1},"<>",H6:H{row-1})',
         None,
         f'=SUMIF(E6:E{row-1},"<>",J6:J{row-1})',
         None, None],
        col_start=2, col_end=12)
    ws.cell(row=row, column=2).alignment = mk_align("left")
    ws.cell(row=row, column=2+6).number_format = FMT_INT
    ws.cell(row=row, column=2+8).number_format = FMT_COST
    row += 2

    write_note(ws, row,
        f'数据来源：{os.path.basename(FILE_CURR)}  |  统计时间：{ts_curr}',
        col_end=12)

def _write_s2_subtotal(ws, row, level, ls, le):
    ws.row_dimensions[row].height = 22
    for col in range(2, 13):
        c = ws.cell(row=row, column=col)
        c.fill = mk_fill(C_TOTAL_BG)
        c.font = mk_font(bold=True, size=10)
        c.alignment = mk_align("center")
        c.border = mk_border_medium_top()
    c = ws.cell(row=row, column=2, value=f'小计')
    c.alignment = mk_align("left")
    # SKU数量
    ws.cell(row=row, column=2+2,  value=f'=COUNTA(C{ls}:C{le})')
    # 总件数
    ws.cell(row=row, column=2+6,  value=f'=SUM(H{ls}:H{le})')
    ws.cell(row=row, column=2+6).number_format = FMT_INT
    # 总成本
    ws.cell(row=row, column=2+8,  value=f'=SUM(J{ls}:J{le})')
    ws.cell(row=row, column=2+8).number_format = FMT_COST


# ═══════════════════════════════════════════════════════════════
# SHEET 3: 库存变动对比（仅高流通 + 有变动）
# ═══════════════════════════════════════════════════════════════
def make_s3_cols(ts_prev, ts_curr):
    return ['品类', 'SKU名称', '系列', '库存级别',
            f'上次件数\n({ts_prev})', f'上次成本\n({ts_prev})',
            f'本次件数\n({ts_curr})', f'本次成本\n({ts_curr})',
            '变动件数', '变动成本']

def build_sheet3(wb, df_curr, df_prev, ts_curr, ts_prev):
    ws = wb.create_sheet('库存变动对比')
    ws.sheet_properties.tabColor = "3D6B4F"
    set_no_gridlines(ws)
    set_col_widths(ws, {
        'B': 12, 'C': 32, 'D': 14, 'E': 14,
        'F': 9, 'G': 11, 'H': 9, 'I': 11,
        'J': 9, 'K': 11,
    })

    # ── 先计算变动数据，再写摘要 ──────────────────────────────
    def mk_qty(r):
        return to_num(r.get('2026 滨江仓库数量', 0)) + to_num(r.get('2026 常州数量', 0))

    prev_map = {}
    for _, r in df_prev.iterrows():
        name = normalize_sku(str(r.get('名称', '') or ''))
        prev_map[name] = {
            'qty': mk_qty(r),
            'cost': to_num(r.get('实际成本价', 0)),
        }

    changed_rows = []
    for _, r in df_curr.iterrows():
        lvl = str(r.get('库存级别', '') or '').strip()
        if lvl not in HIGH_CIRC_LEVELS:
            continue
        name = normalize_sku(str(r.get('名称', '') or ''))
        curr_qty = mk_qty(r)
        curr_cost = to_num(r.get('实际成本价', 0))
        prev_data = prev_map.get(name, {'qty': 0, 'cost': curr_cost})
        prev_qty = prev_data['qty']
        if curr_qty == prev_qty:
            continue
        # 若旧成本价为0，统一使用当前成本价，避免虚假成本变动
        prev_cost = prev_data['cost'] if to_num(prev_data['cost']) > 0 else curr_cost
        changed_rows.append({
            'row_data': r,
            'prev_qty': prev_qty,
            'prev_cost': prev_cost,
            'curr_qty': curr_qty,
            'curr_cost': curr_cost,
            'delta_qty': curr_qty - prev_qty,
            'lvl': lvl,
        })

    def sort_key(x):
        r = x['row_data']
        return (
            level_idx(x['lvl']),
            cat_idx(str(r.get('品类', '') or '')),
            str(r.get('系列', '') or '').strip().lower(),
            str(r.get('名称', '') or '').strip().lower(),
        )
    changed_rows.sort(key=sort_key)

    # 摘要统计
    out_cnt = sum(1 for x in changed_rows if x['delta_qty'] < 0)
    in_cnt  = sum(1 for x in changed_rows if x['delta_qty'] > 0)
    total_delta_qty  = sum(x['delta_qty'] for x in changed_rows)
    total_prev_cost  = sum(x['prev_qty'] * x['prev_cost'] for x in changed_rows)
    total_curr_cost  = sum(x['curr_qty'] * x['curr_cost'] for x in changed_rows)
    total_delta_cost = total_curr_cost - total_prev_cost

    in_note  = f"入库 {in_cnt} 个SKU" if in_cnt > 0 else "无入库记录"
    summary_line = (
        f"本期（{ts_prev} → {ts_curr}）高流通库存共 {len(changed_rows)} 个SKU发生变动："
        f"出库 {out_cnt} 个SKU，{in_note}。"
        f"合计出库 {abs(total_delta_qty):.0f} 件，"
        f"库存成本减少 ¥{abs(total_delta_cost):,.0f}。"
    )

    # ── 写标题 + 摘要 + 列标题 ──────────────────────────────
    write_title_block(ws, 2,
        f'库存变动对比  ·  {ts_prev} → {ts_curr}  ·  高流通品类',
        '仅含高流通级别（正常商品/DA·Zipp/换货回/FSA/自营/套件/配件/铺货）· 仅显示有变动的SKU',
        col_end=11)

    # 摘要文字行（row 5）
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=11)
    c = ws.cell(row=5, column=2, value=summary_line)
    c.fill = mk_fill(C_PAPER)
    c.font = mk_font(size=10, bold=True, color=C_NEAR_BLACK)
    c.alignment = mk_align("left")
    ws.row_dimensions[5].height = 20

    # KPI 汇总行（row 6）：上次总成本 | 本次总成本 | 变动成本
    ws.row_dimensions[6].height = 22
    kpi_labels = [
        (f'上次总成本 ({ts_prev})', total_prev_cost, FMT_COST),
        (f'本次总成本 ({ts_curr})', total_curr_cost, FMT_COST),
        ('成本变动',          total_delta_cost, FMT_COST),
        ('合计变动件数',       total_delta_qty,  FMT_INT),
    ]
    kpi_col = 2
    for label, val, fmt in kpi_labels:
        # label cell
        cl = ws.cell(row=6, column=kpi_col, value=label)
        cl.fill = mk_fill(C_NEAR_BLACK); cl.font = mk_font(bold=True, size=9, color=C_WHITE)
        cl.alignment = mk_align("center")
        kpi_col += 1
        # value cell
        cv = ws.cell(row=6, column=kpi_col, value=val)
        cv.fill = mk_fill(C_TOTAL_BG); cv.font = mk_font(bold=True, size=10)
        cv.number_format = fmt; cv.alignment = mk_align("center")
        # 变动值用勃艮第红
        if label in ('成本变动', '合计变动件数') and val < 0:
            cv.font = mk_font(bold=True, size=10, color=C_BURGUNDY)
        kpi_col += 1
    # 填满剩余列
    for col in range(kpi_col, 12):
        ws.cell(row=6, column=col).fill = mk_fill(C_TOTAL_BG)

    # 分隔空行（row 7）
    ws.row_dimensions[7].height = 6
    for col in range(2, 12):
        ws.cell(row=7, column=col).fill = mk_fill(C_WHITE)

    write_col_headers(ws, 8, make_s3_cols(ts_prev, ts_curr), col_start=2)

    row = 9
    prev_level = None
    level_start = {}
    level_end = {}

    for item in changed_rows:
        r = item['row_data']
        lvl = item['lvl']

        if lvl != prev_level:
            if prev_level is not None:
                level_end[prev_level] = row - 1
                _write_s3_subtotal(ws, row, prev_level, level_start[prev_level], row-1)
                row += 1
            icon = LEVEL_ICONS.get(lvl, '⚪')
            write_level_header(ws, row, f'{icon} {lvl}', col_end=11)
            row += 1
            level_start[lvl] = row
            prev_level = lvl

        stripe = (row % 2 == 0)
        bg = C_ROW_STRIPE if stripe else C_WHITE
        ws.row_dimensions[row].height = 22

        prev_cost_total = item['prev_qty'] * item['prev_cost']
        curr_cost_total = item['curr_qty'] * item['curr_cost']

        vals = [
            str(r.get('品类', '') or '').strip(),
            str(r.get('名称', '') or '').strip(),
            str(r.get('系列', '') or '').strip(),
            lvl,
            item['prev_qty'],
            prev_cost_total,
            item['curr_qty'],
            curr_cost_total,
            f'=H{row}-F{row}',     # 变动件数
            f'=I{row}-G{row}',     # 变动成本
        ]

        for i, val in enumerate(vals):
            col = 2 + i
            c = ws.cell(row=row, column=col, value=val)
            c.fill = mk_fill(bg)
            c.font = mk_font(size=10)
            c.border = mk_border_thin()
            c.alignment = mk_align("left") if i == 1 else mk_align("center")

        # 数字格式
        for col_offset, fmt in [(4,FMT_INT),(5,FMT_COST),(6,FMT_INT),(7,FMT_COST),(8,FMT_INT),(9,FMT_COST)]:
            ws.cell(row=row, column=2+col_offset).number_format = fmt

        # 负变动用勃艮第红
        delta = item['delta_qty']
        if delta < 0:
            ws.cell(row=row, column=2+8).font = mk_font(size=10, color=C_BURGUNDY, bold=True)
            ws.cell(row=row, column=2+9).font = mk_font(size=10, color=C_BURGUNDY, bold=True)

        row += 1

    if prev_level:
        level_end[prev_level] = row - 1
        _write_s3_subtotal(ws, row, prev_level, level_start[prev_level], row-1)
        row += 1

    # 变动合计
    ws.row_dimensions[row].height = 22
    write_subtotal_row(ws, row,
        ['变动合计', None, None, None,
         f'=SUM(F9:F{row-1})', f'=SUM(G9:G{row-1})',
         f'=SUM(H9:H{row-1})', f'=SUM(I9:I{row-1})',
         f'=SUM(J9:J{row-1})', f'=SUM(K9:K{row-1})'],
        col_start=2, col_end=11)
    ws.cell(row=row, column=2).alignment = mk_align("left")
    for col_offset, fmt in [(4,FMT_INT),(5,FMT_COST),(6,FMT_INT),(7,FMT_COST),(8,FMT_INT),(9,FMT_COST)]:
        ws.cell(row=row, column=2+col_offset).number_format = fmt
    row += 2

    write_note(ws, row,
        f'数据来源：{os.path.basename(FILE_PREV)} vs {os.path.basename(FILE_CURR)}  |  统计时间：{ts_curr}',
        col_end=11)

def _write_s3_subtotal(ws, row, level, ls, le):
    ws.row_dimensions[row].height = 22
    for col in range(2, 12):
        c = ws.cell(row=row, column=col)
        c.fill = mk_fill(C_TOTAL_BG)
        c.font = mk_font(bold=True, size=10)
        c.alignment = mk_align("center")
        c.border = mk_border_medium_top()
    c = ws.cell(row=row, column=2, value='小计')
    c.alignment = mk_align("left")
    ws.cell(row=row, column=2+4).value = f'=SUM(F{ls}:F{le})'
    ws.cell(row=row, column=2+4).number_format = FMT_INT
    ws.cell(row=row, column=2+5).value = f'=SUM(G{ls}:G{le})'
    ws.cell(row=row, column=2+5).number_format = FMT_COST
    ws.cell(row=row, column=2+6).value = f'=SUM(H{ls}:H{le})'
    ws.cell(row=row, column=2+6).number_format = FMT_INT
    ws.cell(row=row, column=2+7).value = f'=SUM(I{ls}:I{le})'
    ws.cell(row=row, column=2+7).number_format = FMT_COST
    ws.cell(row=row, column=2+8).value = f'=SUM(J{ls}:J{le})'
    ws.cell(row=row, column=2+8).number_format = FMT_INT
    ws.cell(row=row, column=2+9).value = f'=SUM(K{ls}:K{le})'
    ws.cell(row=row, column=2+9).number_format = FMT_COST


# ═══════════════════════════════════════════════════════════════
# SHEET 4: 清理库存跟踪  (含期初列)
# ═══════════════════════════════════════════════════════════════
# 列: B=SKU  C=品类  D=系列
#     E=期初件  F=期初成本
#     G=3/12件  H=3/12成本
#     I=3/13件  J=3/13成本
#     K=已清件(=E-I)  L=已清成本(=F-J)
#     M=4/15目标  N=4/15实际  O=完成率  P=6/15目标  Q=备注
def make_s4_cols(ts_prev, ts_curr, ts_qichu):
    return [
        'SKU名称', '品类', '系列',
        f'期初\n件数\n({ts_qichu})', f'期初\n总成本\n({ts_qichu})',
        f'{ts_prev}\n件数', f'{ts_prev}\n成本',
        f'{ts_curr}\n件数', f'{ts_curr}\n成本',
        '已清\n件数', '已清\n成本',
        '备注',
        '4/15\n清货目标', '完成率', '6/15\n清货目标',
    ]

S4_COL_END = 16   # B=2 .. P=16
# N/O/P 列仅在人员小计行填写，SKU数据行留空

PERSONS = ['张振', '郭城', '潘昊', '海尔', '程总', '(待确认)']

# 清货目标金额（手动维护，每次调整在此处修改即可）
# 格式：'人名': (4/15清货目标, 6/15清货目标)，None 表示未设定
PERSON_TARGETS = {
    '张振':   (250_000, 450_000),
    '郭城':   (150_000, 250_000),
    '潘昊':   ( 70_000, 100_000),
    '海尔':   (100_000,    None),
    '程总':   (   None,    None),
}


def build_sheet4(wb, df_curr, df_prev, person_map, qichu_data, ts_curr, ts_prev, ts_qichu):
    ws = wb.create_sheet('清理库存跟踪')
    ws.sheet_properties.tabColor = C_BURGUNDY
    ws.sheet_properties.outlinePr.summaryBelow = True   # 小计在组下方
    set_no_gridlines(ws)
    set_col_widths(ws, {
        'B': 30, 'C': 10, 'D': 13,
        'E': 8,  'F': 11,
        'G': 8,  'H': 11,
        'I': 8,  'J': 11,
        'K': 8,  'L': 11,
        'M': 18,
        'N': 13, 'O': 10, 'P': 13,
    })

    write_title_block(ws, 2,
        f'清理库存跟踪  ·  {ts_curr}',
        f'期初({ts_qichu}) → {ts_prev} → {ts_curr}  ·  含 Pinarello 清理 + 清理库存  ·  按负责人分组  ·  M/P列目标请手动填写',
        col_end=S4_COL_END)

    # 使用说明行
    usage = (f'已清成本 = 期初成本 - 当前成本  ·  完成率 = 已清成本 / 4/15清货目标  ·  N/P列目标为人工手填金额')
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=S4_COL_END)
    c = ws.cell(row=5, column=2, value=usage)
    c.fill = mk_fill(C_PAPER); c.font = mk_font(size=9, color=C_MUTED)
    c.alignment = mk_align("left"); ws.row_dimensions[5].height = 16

    write_col_headers(ws, 6, make_s4_cols(ts_prev, ts_curr, ts_qichu), col_start=2)

    # 获取清理级别的SKU
    clearance_levels = {'Pinarello 清理', '清理库存'}
    df_clear = df_curr[df_curr['库存级别'].isin(clearance_levels)].copy()

    def mk_qty(r_dict):
        return to_num(r_dict.get('2026 滨江仓库数量', 0)) + to_num(r_dict.get('2026 常州数量', 0))

    # 3/12 数据字典
    prev_map = {}
    for _, r in df_prev.iterrows():
        name = normalize_sku(str(r.get('名称', '') or ''))
        prev_map[name] = {
            'qty': mk_qty(r),
            'cost_price': to_num(r.get('实际成本价', 0)),
        }

    # 分组到各人（按品类→成本降序排列）
    person_skus = {p: [] for p in PERSONS}
    for _, r in df_clear.iterrows():
        name_key = normalize_sku(str(r.get('名称', '') or ''))
        matched_person = person_map.get(name_key, '(待确认)')
        if matched_person not in PERSONS:
            matched_person = '(待确认)'
        person_skus[matched_person].append(r)

    # 每个人的SKU按品类→系列→SKU名称排序
    def sort_person_skus(skus):
        return sorted(skus, key=lambda r: (
            cat_idx(str(r.get('品类', '') or '')),
            str(r.get('系列', '') or '').strip().lower(),
            str(r.get('名称', '') or '').strip().lower(),
        ))

    row = 7

    for person in PERSONS:
        skus = sort_person_skus(person_skus[person])
        if not skus:
            continue

        if person == '张振':
            pina = []
            non_pina = []
            for r in skus:
                series = str(r.get('系列', '') or '').strip()
                name   = str(r.get('名称', '') or '').strip()
                if series.startswith('Pinarello') or name.startswith('Dogma'):
                    pina.append(r)
                else:
                    non_pina.append(r)

            _write_person_header(ws, row, person, S4_COL_END); row += 1

            pina_subtotal_row = None
            np_subtotal_row   = None

            if pina:
                _write_subgroup_header(ws, row, '  ▸ Pinarello / Dogma 清理', S4_COL_END, color=C_BURGUNDY)
                row += 1
                pina_start = row
                for idx, r_data in enumerate(pina):
                    row = _write_s4_data_row(ws, row, r_data, prev_map, qichu_data, idx, outline_level=2)
                pina_subtotal_row = row
                _write_person_subtotal(ws, row, '  Pinarello 小计', pina_start, row-1, S4_COL_END)
                ws.row_dimensions[row].outlineLevel = 1
                row += 1

            if non_pina:
                _write_subgroup_header(ws, row, '  ▸ 非 Pinarello', S4_COL_END)
                row += 1
                np_start = row
                for idx, r_data in enumerate(non_pina):
                    row = _write_s4_data_row(ws, row, r_data, prev_map, qichu_data, idx, outline_level=2)
                np_subtotal_row = row
                _write_person_subtotal(ws, row, '  非 Pinarello 小计', np_start, row-1, S4_COL_END)
                ws.row_dimensions[row].outlineLevel = 1
                row += 1

            # 张振总合计行（汇总两子组，承载目标 KPI）
            _write_zz_total(ws, row, '张振 小计', pina_subtotal_row, np_subtotal_row, S4_COL_END,
                            targets=PERSON_TARGETS.get('张振', (None, None)))
            row += 1

        else:
            _write_person_header(ws, row, person, S4_COL_END); row += 1
            data_start = row
            for idx, r_data in enumerate(skus):
                row = _write_s4_data_row(ws, row, r_data, prev_map, qichu_data, idx, outline_level=1)
            _write_person_subtotal(ws, row, f'{person} 小计', data_start, row-1, S4_COL_END,
                                   show_targets=True, targets=PERSON_TARGETS.get(person, (None, None)))
            row += 1

        row += 1  # 人员间空行

    # 总合计（汇总所有人期初→当前→已清，不含人员目标列）
    row += 1
    write_subtotal_row(ws, row,
        ['合计', None, None,
         f'=SUM(E7:E{row-1})', f'=SUM(F7:F{row-1})',
         f'=SUM(G7:G{row-1})', f'=SUM(H7:H{row-1})',
         f'=SUM(I7:I{row-1})', f'=SUM(J7:J{row-1})',
         f'=IF(SUMPRODUCT((E7:E{row-1})-(I7:I{row-1}))=0,"",SUM(K7:K{row-1}))',
         f'=IF(SUMPRODUCT((F7:F{row-1})-(J7:J{row-1}))=0,"",SUM(L7:L{row-1}))',
         None, None, None, None],
        col_start=2, col_end=S4_COL_END)
    ws.cell(row=row, column=2).alignment = mk_align("left")
    for off, fmt in [(3,FMT_INT),(4,FMT_COST),(5,FMT_INT),(5,FMT_COST),
                     (6,FMT_INT),(7,FMT_COST),(8,FMT_INT),(9,FMT_COST),
                     (10,FMT_INT),(11,FMT_COST)]:
        ws.cell(row=row, column=2+off).number_format = fmt
    row += 2

    write_note(ws, row,
        f'期初数据来源：2026清货KPI管理表_表1表2（{ts_qichu}）· 上次来源：{os.path.basename(FILE_PREV)} · 本次来源：{os.path.basename(FILE_CURR)}',
        col_end=S4_COL_END)


def _write_subgroup_header(ws, row, label, col_end, color=C_NEAR_BLACK):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=2, value=label)
    c.fill = mk_fill(C_ROW_STRIPE)
    c.font = mk_font(bold=True, size=10, color=color)
    c.alignment = mk_align("left"); ws.row_dimensions[row].height = 20


def _write_person_header(ws, row, person, col_end):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=2, value=f'▸  {person}')
    c.fill = mk_fill(C_NEAR_BLACK)
    c.font = mk_font(bold=True, size=11, color=C_WHITE)
    c.alignment = mk_align("left"); ws.row_dimensions[row].height = 26


def _write_s4_data_row(ws, row, r_data, prev_map, qichu_data, idx, outline_level=1):
    """写Sheet4数据行（含期初列）"""
    name_orig = str(r_data.get('名称', '') or '').strip()
    name_key  = normalize_sku(name_orig)

    # 3/13 数据（先算，供下方回退使用）
    curr_qty        = to_num(r_data.get('2026 滨江仓库数量', 0)) + to_num(r_data.get('2026 常州数量', 0))
    curr_cost_price = to_num(r_data.get('实际成本价', 0))
    curr_cost       = curr_qty * curr_cost_price

    # 3/12 数据：若旧成本价为0，回退到当前成本价，避免虚假变动
    prev = prev_map.get(name_key, {})
    prev_qty   = prev.get('qty', None)
    raw_prev_cost_price = to_num(prev.get('cost_price', 0))
    prev_cost_price = raw_prev_cost_price if raw_prev_cost_price > 0 else curr_cost_price
    prev_cost  = (prev_qty * prev_cost_price) if prev_qty is not None else None
    # 3/12 无记录但当前有库存 → 视为一直存在
    if prev_qty is None and curr_qty > 0:
        prev_qty  = curr_qty
        prev_cost = curr_cost

    # 期初数据
    qc = qichu_data.get(name_key, {})
    qichu_qty   = qc.get('qty', None)
    qichu_cost  = qc.get('total_cost', None)
    # 期初有件数但成本为0（当时未录入）→ 用当前成本价回填
    if qichu_qty and to_num(qichu_cost) == 0 and curr_cost_price > 0:
        qichu_cost = qichu_qty * curr_cost_price
    # 期初无记录但当前有库存 → 视为一开始就有、未发生清货，期初 = 当前
    if qichu_qty is None and curr_qty > 0:
        qichu_qty  = curr_qty
        qichu_cost = curr_cost

    # 已清公式：期初 - 当前(3/13)，为0时显示空白
    cleared_qty_f  = f'=IF(E{row}-I{row}=0,"",E{row}-I{row})' if qichu_qty is not None else '—'
    cleared_cost_f = f'=IF(F{row}-J{row}=0,"",F{row}-J{row})' if qichu_cost is not None else '—'

    stripe = (idx % 2 == 1)
    bg = C_ROW_STRIPE if stripe else C_WHITE
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row].outlineLevel = outline_level   # 可折叠分组

    vals = [
        name_orig,                          # B SKU名称
        str(r_data.get('品类', '') or '').strip(),   # C
        str(r_data.get('系列', '') or '').strip(),   # D
        qichu_qty,                          # E 期初件数
        qichu_cost,                         # F 期初成本
        prev_qty,                           # G 3/12件数
        prev_cost,                          # H 3/12成本
        curr_qty,                           # I 3/13件数
        curr_cost,                          # J 3/13成本
        cleared_qty_f,                      # K 已清件数
        cleared_cost_f,                     # L 已清成本
        str(r_data.get('其他销售备注', '') or '').strip(),  # M 备注
    ]

    num_fmts = {
        3: FMT_INT, 4: FMT_COST,
        5: FMT_INT, 6: FMT_COST,
        7: FMT_INT, 8: FMT_COST,
        9: FMT_INT, 10: FMT_COST,
    }

    for i, val in enumerate(vals):
        col = 2 + i
        c = ws.cell(row=row, column=col, value=val)
        c.fill = mk_fill(bg)
        c.font = mk_font(size=10)
        c.border = mk_border_thin()
        c.alignment = mk_align("left") if i in (0, 11) else mk_align("center")
        if i in num_fmts:
            c.number_format = num_fmts[i]

    return row + 1


def _write_zz_total(ws, row, label, pina_row, np_row, col_end, targets=(None, None)):
    """张振总合计行：汇总 Pinarello 和 非Pinarello 两个子组的小计行，并提供目标 KPI。"""
    ws.row_dimensions[row].height = 22
    for col in range(2, col_end+1):
        c = ws.cell(row=row, column=col)
        c.fill = mk_fill(C_TOTAL_BG)
        c.font = mk_font(bold=True, size=10)
        c.alignment = mk_align("center")
        c.border = mk_border_medium_top()
    ws.cell(row=row, column=2, value=label).alignment = mk_align("left")

    # 汇总两子组：用 N() 函数把 "" 当 0 处理，避免对空字符串求和出错
    refs = []
    if pina_row: refs.append(pina_row)
    if np_row:   refs.append(np_row)

    for col_letter, offset, fmt in [
        ('E', 3, FMT_INT), ('F', 4, FMT_COST),
        ('G', 5, FMT_INT), ('H', 6, FMT_COST),
        ('I', 7, FMT_INT), ('J', 8, FMT_COST),
    ]:
        formula = '+'.join(f'{col_letter}{r}' for r in refs)
        c = ws.cell(row=row, column=2+offset, value=f'={formula}')
        c.number_format = fmt

    for col_letter, offset, fmt in [('K', 9, FMT_INT), ('L', 10, FMT_COST)]:
        parts = '+'.join(f'N({col_letter}{r})' for r in refs)
        c = ws.cell(row=row, column=2+offset,
                    value=f'=IF({parts}=0,"",{parts})')
        c.number_format = fmt

    t415, t615 = targets
    # N: 4/15清货目标
    c_n = ws.cell(row=row, column=2+12, value=t415)
    c_n.fill = mk_fill(C_AMBER); c_n.border = mk_border_medium_top()
    c_n.font = mk_font(bold=True, size=10); c_n.number_format = FMT_COST
    # O: 完成率
    c_o = ws.cell(row=row, column=2+13,
                  value=f'=IF(OR(N{row}="",L{row}=""),"",L{row}/N{row})')
    c_o.number_format = FMT_PCT; c_o.border = mk_border_medium_top()
    c_o.font = mk_font(bold=True, size=10)
    # P: 6/15清货目标
    c_p = ws.cell(row=row, column=2+14, value=t615)
    c_p.fill = mk_fill(C_AMBER); c_p.border = mk_border_medium_top()
    c_p.font = mk_font(bold=True, size=10); c_p.number_format = FMT_COST


def _write_person_subtotal(ws, row, label, data_start, data_end, col_end, show_targets=False, targets=(None, None)):
    """
    写人员/子组小计行。
    show_targets=True → 在 N(4/15目标)/O(完成率)/P(6/15目标) 列额外写入人员级 KPI。
    """
    ws.row_dimensions[row].height = 22
    for col in range(2, col_end+1):
        c = ws.cell(row=row, column=col)
        c.fill = mk_fill(C_TOTAL_BG)
        c.font = mk_font(bold=True, size=10)
        c.alignment = mk_align("center")
        c.border = mk_border_medium_top()
    c = ws.cell(row=row, column=2, value=label)
    c.alignment = mk_align("left")

    # B=SKU名称 C=品类 D=系列 E=期初件数 F=期初成本 G=3/12件 H=3/12成本
    # I=3/13件 J=3/13成本 K=已清件 L=已清成本 M=备注 N=4/15清货目标 O=完成率 P=6/15清货目标
    formulas = [
        (3,  f'=SUM(E{data_start}:E{data_end})',  FMT_INT),   # E 期初件数
        (4,  f'=SUM(F{data_start}:F{data_end})',  FMT_COST),  # F 期初成本
        (5,  f'=SUM(G{data_start}:G{data_end})',  FMT_INT),   # G 3/12件数
        (6,  f'=SUM(H{data_start}:H{data_end})',  FMT_COST),  # H 3/12成本
        (7,  f'=SUM(I{data_start}:I{data_end})',  FMT_INT),   # I 3/13件数
        (8,  f'=SUM(J{data_start}:J{data_end})',  FMT_COST),  # J 3/13成本
        (9,  f'=IF(SUMPRODUCT((E{data_start}:E{data_end})-(I{data_start}:I{data_end}))=0,"",SUM(K{data_start}:K{data_end}))',  FMT_INT),
        (10, f'=IF(SUMPRODUCT((F{data_start}:F{data_end})-(J{data_start}:J{data_end}))=0,"",SUM(L{data_start}:L{data_end}))',  FMT_COST),
    ]
    for off, formula, fmt in formulas:
        c = ws.cell(row=row, column=2+off, value=formula)
        c.number_format = fmt

    if show_targets:
        t415, t615 = targets
        # N: 4/15清货目标
        c_n = ws.cell(row=row, column=2+12, value=t415)
        c_n.fill = mk_fill(C_AMBER)
        c_n.border = mk_border_medium_top()
        c_n.font = mk_font(bold=True, size=10)
        c_n.number_format = FMT_COST
        # O: 完成率 = 已清成本(L) / 4/15目标(N)
        c_o = ws.cell(row=row, column=2+13,
                      value=f'=IF(OR(N{row}="",L{row}=""),"",L{row}/N{row})')
        c_o.number_format = FMT_PCT
        c_o.border = mk_border_medium_top()
        c_o.font = mk_font(bold=True, size=10)
        # P: 6/15清货目标
        c_p = ws.cell(row=row, column=2+14, value=t615)
        c_p.fill = mk_fill(C_AMBER)
        c_p.border = mk_border_medium_top()
        c_p.font = mk_font(bold=True, size=10)
        c_p.number_format = FMT_COST


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    print("📊 读取数据...")
    df_curr = load_df(FILE_CURR)
    df_prev = load_df(FILE_PREV)
    print(f"  {os.path.basename(FILE_CURR)}: {len(df_curr)} 行")
    print(f"  {os.path.basename(FILE_PREV)}: {len(df_prev)} 行")

    # ── 品类级别修正规则 ──────────────────────────────────────────
    # 头盔全部归入「配件」级别（清理类不动）
    CLEARANCE_LEVELS = {'Pinarello 清理', '清理库存'}
    CATEGORY_LEVEL_OVERRIDES = {
        '头盔': '配件',   # 头盔统一归配件，不区分正常/特殊
    }
    def apply_category_overrides(df):
        df = df.copy()
        for cat, target_level in CATEGORY_LEVEL_OVERRIDES.items():
            mask = (df['品类'] == cat) & (~df['库存级别'].isin(CLEARANCE_LEVELS))
            changed = mask.sum()
            if changed:
                print(f"  品类修正：{cat} → {target_level}（{changed} 条）")
            df.loc[mask, '库存级别'] = target_level
        return df

    df_curr = apply_category_overrides(df_curr)
    df_prev = apply_category_overrides(df_prev)

    print("👤 读取 v6 人员分配 + 手动覆盖...")
    person_map = load_v6_person_map()
    overrides = {
        # 原 (待确认) 已确认分配 ─── from image
        'selle italia  座垫 3d 碳轨 黑色（散装）':        '海尔',
        'selle italia 座垫 3d 碳轨 黑色（散装）':         '海尔',
        '萨玛仕 座垫 3d打印（电助力用）':                  '海尔',
        'carbon—ti 盘片 55-40':                          '郭城',
        'fsa 轮组 vision 45 rs':                         '郭城',
        'dt 轮组 arc 1100 38 (sh)':                      '郭城',
        'scom 风刃 轮组 49/67 银标 碟刹':                 '潘昊',
        'scom 风刃 轮组 49/67 黑标 碟刹':                 '潘昊',
        'pcw 轮组 古铜金6560':                            '海尔',
        'gusto 整车 gtr 白色 m':                          '郭城',
        'cline 整车 火焰清漆m2':                          '郭城',
        # 张振确认
        'dogma x 车架 bob 515':                          '张振',
        'pinarello 整车 f9 宝石蓝 515':                  '张振',
        # DEDA 80%郭城 / 20%张振 (按件数)
        # 郭城: 40-100(15件) + 40-110(11件) + 44-100(9件) + 44-120(8件) = 43件 ≈ 83%
        'deda 车把一体把 alanera rs 40-100':              '郭城',
        'deda 车把一体把 alanera rs 40-110':              '郭城',
        'deda 车把一体把 alanera rs 44-100':              '郭城',
        'deda 车把一体把 alanera rs 44-120':              '郭城',
        # 张振: 44-110(7件) + 42-100(1件) + 42-110(1件) = 9件 ≈ 17%
        'deda 车把一体把 alanera rs 44-110':              '张振',
        'deda 车把一体把 alanera rs 42-100':              '张振',
        'deda 车把一体把 alanera rs 42-110':              '张振',
    }
    person_map.update(overrides)
    print(f"  人员映射: {len(person_map)} 条 SKU（含 {len(overrides)} 条手动覆盖）")

    print("📅 读取期初(2/8)数据...")
    qichu_data = load_qichu_data()
    print(f"  期初数据: {len(qichu_data)} 条 SKU")

    # 解析文件名时间戳
    ts_curr, dt_curr = parse_ts(FILE_CURR)   # e.g. '3/13 14:00', '3/13'
    ts_prev, dt_prev = parse_ts(FILE_PREV)   # e.g. '3/12 14:00', '3/12'
    TS_QICHU = '2/8'                          # 期初无具体时间
    print(f"  时间戳  期初={TS_QICHU}  上次={ts_prev}  本次={ts_curr}")

    # Sheet1/2 只展示有在库件数的 SKU；Sheet4 清理跟踪保留零库存（显示已清进度）
    def get_qty_series(df):
        bj = df['2026 滨江仓库数量'].apply(to_num)
        cz = df['2026 常州数量'].apply(to_num)
        return bj + cz

    df_instock = df_curr[get_qty_series(df_curr) > 0].copy()
    print(f"  有库存 SKU: {len(df_instock)} 条（过滤掉 {len(df_curr)-len(df_instock)} 条零库存）")

    print("📝 构建工作簿...")
    wb = openpyxl.Workbook()

    print("  Sheet 1: 总览")
    build_sheet1(wb, df_instock, ts_curr)

    print("  Sheet 2: 库存SKU明细")
    build_sheet2(wb, df_instock, ts_curr)

    print("  Sheet 3: 库存变动对比")
    build_sheet3(wb, df_curr, df_prev, ts_curr, ts_prev)

    print("  Sheet 4: 清理库存跟踪")
    build_sheet4(wb, df_curr, df_prev, person_map, qichu_data, ts_curr, ts_prev, TS_QICHU)

    print(f"💾 保存到 {OUT}")
    wb.save(OUT)
    print("✅ 完成！")

    # 校验
    wb2 = openpyxl.load_workbook(OUT, data_only=False)
    for s in wb2.sheetnames:
        ws = wb2[s]
        formula_cnt = sum(
            1 for row in ws.iter_rows()
            for cell in row
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('=')
        )
        print(f"  {s}: {ws.max_row}行 x {ws.max_column}列, {formula_cnt}个公式")

    # ── 数据质量审核 ─────────────────────────────────────────────────────
    run_audit(df_curr, df_prev, qichu_data, person_map)


# ═══════════════════════════════════════════════════════════════════════
# 数据质量审核函数（每次生成报表后自动执行）
# ═══════════════════════════════════════════════════════════════════════
def run_audit(df_curr, df_prev, qichu_data, person_map):
    """
    扫描数据质量问题并打印审核报告。
    分三类：🔴 必须修复 / 🟡 建议确认 / 🔵 信息提示
    """
    issues_red    = []
    issues_yellow = []
    issues_blue   = []

    clearance_levels = {'Pinarello 清理', '清理库存'}

    for _, r in df_curr.iterrows():
        name  = str(r.get('名称', '') or '').strip()
        level = str(r.get('库存级别', '') or '').strip()
        cost  = to_num(r.get('实际成本价', 0))
        bj    = to_num(r.get('2026 滨江仓库数量', 0))
        cz    = to_num(r.get('2026 常州数量', 0))
        qty   = bj + cz
        key   = normalize_sku(name)

        # ── 🔴 有库存但成本价为0（最高优先级，会影响所有金额计算）
        if qty > 0 and cost == 0 and level != '铺货商品':
            issues_red.append(f"  成本价=0 | {level:<12} | {name}")

        # ── 🟡 清理SKU未分配负责人
        if level in clearance_levels and key not in person_map:
            issues_yellow.append(f"  未分配负责人 | qty={qty:.0f} | {name}")

        # ── 🟡 期初有记录但成本为0（已触发自动回填，需确认成本价正确）
        qc = qichu_data.get(key, {})
        if qc and to_num(qc.get('total_cost', 0)) == 0 and qc.get('qty', 0) and cost > 0:
            issues_yellow.append(f"  期初成本=0已用当前价回填 ¥{cost:,.0f} | {name}")

        # ── 🔵 3/12成本价为0（已触发回退，但原始数据仍为空）
        prev = df_prev[df_prev['名称'].astype(str).str.strip() == name]
        if not prev.empty:
            prev_cost = to_num(prev.iloc[0].get('实际成本价', 0))
            if prev_cost == 0 and cost > 0 and qty > 0:
                issues_blue.append(f"  3/12成本价=0已回退 → ¥{cost:,.0f} | {name}")

    # 检查当前快照中新增、在上次快照中不存在的清理SKU
    prev_names = set(normalize_sku(str(r.get('名称',''))) for _, r in df_prev.iterrows())
    for _, r in df_curr.iterrows():
        level = str(r.get('库存级别','') or '')
        if level in clearance_levels:
            key = normalize_sku(str(r.get('名称','') or ''))
            if key not in prev_names:
                issues_blue.append(f"  新增清理SKU（上次快照无此记录）| {r.get('名称','')}")

    # ── 输出报告
    sep = "─" * 60
    print(f"\n{'═'*60}")
    print(f"  📋 数据质量审核报告  ·  {os.path.basename(FILE_CURR)}")
    print(f"{'═'*60}")

    if issues_red:
        print(f"\n🔴 必须修复（{len(issues_red)} 项）—— 这些 SKU 在飞书表格中补录成本价后重新导出快照")
        print(sep)
        for x in issues_red:
            print(x)
    else:
        print("\n🔴 必须修复：无")

    if issues_yellow:
        print(f"\n🟡 建议确认（{len(issues_yellow)} 项）—— 数据已自动回填，请人工核实是否正确")
        print(sep)
        for x in issues_yellow:
            print(x)
    else:
        print("🟡 建议确认：无")

    if issues_blue:
        print(f"\n🔵 信息提示（{len(issues_blue)} 项）—— 知悉即可，不影响报表准确性")
        print(sep)
        for x in issues_blue:
            print(x)
    else:
        print("🔵 信息提示：无")

    print(f"\n{'═'*60}")
    if issues_red:
        print("  ⚠️  存在成本价为0的有库存SKU，建议补录后重新生成报表")
    else:
        print("  ✅  数据质量良好，可直接使用本次报表")
    print(f"{'═'*60}\n")


if __name__ == '__main__':
    main()
