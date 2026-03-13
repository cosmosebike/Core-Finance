"""
清理库存对外版生成器
从内部库存报表（vXX.xlsx）提取清理库存，隐去所有成本，保留对外售价，
输出可直接发给渠道分销商/车店的 Excel + HTML 版本。

用法：
  修改 SOURCE_FILE 指向最新版本的内部库存报表，运行即可。
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import math, os, re

# ── 配置 ──────────────────────────────────────────────────
SOURCE_FILE = "库存报表/公路车库存_20260313_v14.xlsx"   # ← 更新数据时改这里
OUT_FILE          = "库存报表/清理库存_对外版_20260313.xlsx"
OUT_HTML          = "库存报表/清理库存_对外版_20260313.html"
OUT_HTML_INTERNAL = "库存报表/清理库存_内部版_20260313.html"

# ── STYLE1 色板（对外版用更简洁的黑白调性）──────────────
C_BLACK    = "111111"
C_HEADER   = "111111"   # 表头黑底
C_WHITE    = "FFFFFF"
C_STRIPE   = "F5F5F5"   # 斑马纹
C_SUBTOTAL = "E8E8E8"   # 小计行
C_TOTAL    = "CCCCCC"   # 合计行
C_BORDER   = "CCCCCC"
C_BURGUNDY = "8B1A1A"   # 仅用于警示/强调价格
C_MUTED    = "555555"   # 次要文字
FONT_NAME  = "PingFang SC"

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def fnt(size=10, bold=False, color=C_BLACK, name=FONT_NAME):
    return Font(name=name, size=size, bold=bold, color=color)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    b = Side(style="medium", color=C_BLACK)
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=b)

def safe_num(x):
    try:
        v = float(x)
        return 0.0 if (math.isnan(v) or math.isinf(v)) else v
    except:
        return 0.0

# ── Step 1：从 Sheet2（库存SKU明细）读取 清理底价 + 成本价 ──
def load_sku_data(wb):
    """返回 {SKU名称: {'清理底价': x, '成本价': x}} 字典"""
    ws = wb.worksheets[1]
    data = {}
    HEADERS = ['品类','SKU名称','系列','库存级别','滨江','常州','总件数','成本价','总成本','清理底价','处理方式']
    for row in ws.iter_rows(min_row=6, values_only=True):
        vals = row[1:]
        if vals[0] is None or vals[1] is None:
            continue
        if str(vals[0]).startswith(('🔵','🔴','⚪','▶','合计','小计','品类')):
            continue
        d = dict(zip(HEADERS, vals))
        sku = str(d.get('SKU名称') or '').strip()
        if sku:
            data[sku] = {
                '清理底价': safe_num(d.get('清理底价')),
                '成本价':   safe_num(d.get('成本价')),
            }
    return data

def load_clearance_price(wb):
    """兼容旧接口：只返回 {SKU名称: 清理底价}"""
    return {k: v['清理底价'] for k, v in load_sku_data(wb).items()}

# ── Step 2：从 Sheet4（清理库存跟踪）读取清理 SKU 列表 ───
def load_clearance_skus(wb):
    """返回清理 SKU 列表：[{SKU名称, 品类, 系列, 当前数量, 备注}]"""
    ws = wb.worksheets[3]
    skus = []
    # 行6 = 表头，行7+ = 数据（含分组行）
    # 列映射：B=SKU名称, C=品类, D=系列, I=3/13件数(当前), M=备注
    COL = {'sku': 1, '品类': 2, '系列': 3, '当前数量': 8, '备注': 12}
    for row in ws.iter_rows(min_row=9, values_only=True):
        sku_val  = row[COL['sku']]
        cat_val  = row[COL['品类']]
        ser_val  = row[COL['系列']]
        qty_val  = row[COL['当前数量']]
        note_val = row[COL['备注']]

        if sku_val is None:
            continue
        sku_str = str(sku_val).strip()
        # 跳过分组标题行（▸ 开头 或 小计/合计）
        if sku_str.startswith(('▸','▶','小计','合计','Pinarello 小计','非 Pinarello','张振','郭城','潘昊','海尔','程总','待确认')):
            continue
        qty = safe_num(qty_val)
        if qty <= 0:
            continue   # 对外只展示有货的

        skus.append({
            'SKU名称':   sku_str,
            '品类':      str(cat_val or '').strip(),
            '系列':      str(ser_val or '').strip(),
            '当前数量':  int(qty),
            '备注':      str(note_val or '').strip(),
        })
    return skus

# ── Step 3：生成 Excel ─────────────────────────────────────
def build_excel(skus, price_map):
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "清理库存报价"
    ws.sheet_view.showGridLines = False

    # 提取时间戳显示用
    import re
    # 文件名尾注格式 MMDDYY（如 031319 = 3月13日19时）
    m2 = re.search(r'_v(\d+)\.xlsx', SOURCE_FILE)
    ts_m = re.search(r'(\d{8})_v', SOURCE_FILE)  # YYYYMMDD
    # 从文件名提取日期，格式：公路车库存_20260313_v14.xlsx
    ts2 = re.search(r'(\d{4})(\d{2})(\d{2})_v', SOURCE_FILE)
    if ts2:
        date_disp = f"{int(ts2.group(2))}月{int(ts2.group(3))}日"
    else:
        date_disp = "3月13日"

    # ── 主标题
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"清理库存 · 渠道报价表"
    c.fill  = fill(C_HEADER)
    c.font  = fnt(14, bold=True, color=C_WHITE)
    c.alignment = aln(h="center")
    ws.row_dimensions[1].height = 36

    # ── 副标题
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"数据截止：{date_disp}  ·  价格为建议清理价，实际以确认为准  ·  库存数量实时变动，请以最终确认为准"
    c.fill  = fill(C_STRIPE)
    c.font  = fnt(9, bold=False, color=C_MUTED)
    c.alignment = aln(h="center")
    ws.row_dimensions[2].height = 22

    # ── 列宽
    col_widths = {'A': 30, 'B': 10, 'C': 18, 'D': 8, 'E': 14, 'F': 10, 'G': 20}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── 表头
    HEADERS = ['SKU 名称', '品类', '系列', '数量', '建议清理价（¥）', '单位', '备注']
    ws.row_dimensions[3].height = 24
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill      = fill(C_HEADER)
        c.font      = fnt(9, bold=True, color=C_WHITE)
        c.alignment = aln(h="center")
        c.border    = border()

    ws.freeze_panes = "A4"

    # ── 按品类 → 系列排序
    CATEGORY_ORDER = ['整车','车架','套件组','牙盘组','轮组','外胎','座垫','锁踏','碟片组','车把一体把','头盔','眼镜','骑行服','其他']
    def sort_key(row):
        cat_pri = CATEGORY_ORDER.index(row['品类']) if row['品类'] in CATEGORY_ORDER else 99
        return (cat_pri, row['系列'], row['SKU名称'])

    skus_sorted = sorted(skus, key=sort_key)

    # 分组写入
    from itertools import groupby
    row_num = 4
    stripe  = False
    ws.sheet_properties.outlinePr.summaryBelow = True

    for cat, cat_group in groupby(skus_sorted, key=lambda x: x['品类']):
        cat_items = list(cat_group)

        # 品类分组标题
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        c = ws.cell(row=row_num, column=1, value=f"  {cat}")
        c.fill  = fill("333333")
        c.font  = fnt(9, bold=True, color=C_WHITE)
        c.alignment = aln(h="left")
        ws.row_dimensions[row_num].height = 20
        group_header = row_num
        row_num += 1
        data_start = row_num

        for item in cat_items:
            price = price_map.get(item['SKU名称'], 0)
            bg = C_STRIPE if stripe else C_WHITE
            stripe = not stripe

            row_vals = [
                (item['SKU名称'], "left"),
                (item['品类'],    "center"),
                (item['系列'],    "center"),
                (item['当前数量'],"center"),
                (price if price > 0 else "—", "right"),
                ("件",            "center"),
                (item['备注'],    "left"),
            ]
            for ci, (val, h) in enumerate(row_vals, 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.fill      = fill(bg)
                c.font      = fnt(9, color=C_BLACK)
                c.alignment = aln(h=h)
                c.border    = border()
                # 价格列：红色强调 + 格式
                if ci == 5 and isinstance(val, (int, float)):
                    c.number_format = '#,##0'
                    c.font = fnt(9, bold=True, color=C_BURGUNDY)
            ws.row_dimensions[row_num].height = 18
            ws.row_dimensions[row_num].outlineLevel = 1
            row_num += 1

        # 品类小计
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        c_st = ws.cell(row=row_num, column=1, value=f"{cat}  小计")
        c_st.fill  = fill(C_SUBTOTAL); c_st.font = fnt(9, bold=True); c_st.border = thick_bottom()
        c_st.alignment = aln(h="left")

        c_q = ws.cell(row=row_num, column=4,
                      value=f"=SUM(D{data_start}:D{row_num-1})")
        c_q.fill  = fill(C_SUBTOTAL); c_q.font = fnt(9, bold=True)
        c_q.alignment = aln(h="center"); c_q.border = thick_bottom()

        for ci in (5, 6, 7):
            c = ws.cell(row=row_num, column=ci)
            c.fill = fill(C_SUBTOTAL); c.border = thick_bottom()

        ws.row_dimensions[row_num].height = 20
        for r_i in range(data_start, row_num):
            ws.row_dimensions[r_i].outlineLevel = 1
        row_num += 1

    # ── 总计行
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    c = ws.cell(row=row_num, column=1, value="总计")
    c.fill = fill(C_TOTAL); c.font = fnt(10, bold=True); c.border = thick_bottom()
    c.alignment = aln(h="left")

    c_total = ws.cell(row=row_num, column=4, value=f"=SUM(D4:D{row_num-1})")
    c_total.fill = fill(C_TOTAL); c_total.font = fnt(10, bold=True)
    c_total.alignment = aln(h="center"); c_total.border = thick_bottom()

    for ci in (5, 6, 7):
        c = ws.cell(row=row_num, column=ci)
        c.fill = fill(C_TOTAL); c.border = thick_bottom()

    ws.row_dimensions[row_num].height = 24

    # ── 底部声明
    row_num += 2
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    c = ws.cell(row=row_num, column=1,
                value="* 以上价格为内部清理建议价，最终成交价以双方确认为准。库存数量实时变动，请在下单前再次确认现货。")
    c.font = fnt(8, color=C_MUTED)
    c.alignment = aln(h="left")
    ws.row_dimensions[row_num].height = 18

    os.makedirs("库存报表", exist_ok=True)
    wb_out.save(OUT_FILE)
    print(f"✅ 已保存：{OUT_FILE}")
    print(f"   SKU 数量：{len(skus)} 条（已过滤零库存）")
    without_price = [s['SKU名称'] for s in skus if price_map.get(s['SKU名称'], 0) == 0]
    if without_price:
        print(f"\n⚠️  以下 {len(without_price)} 个 SKU 未找到清理底价（显示为 —）：")
        for n in without_price[:10]:
            print(f"   {n}")
        if len(without_price) > 10:
            print(f"   ...共 {len(without_price)} 个")

# ── HTML 生成 ─────────────────────────────────────────────
def build_html(skus, price_map):
    ts = re.search(r'(\d{4})(\d{2})(\d{2})_v', SOURCE_FILE)
    date_disp = f"{int(ts.group(2))}月{int(ts.group(3))}日" if ts else "3月13日"
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 排序
    CATEGORY_ORDER = ['整车','车架','套件组','牙盘组','轮组','外胎','座垫','锁踏',
                      '碟片组','车把一体把','头盔','眼镜','骑行服','其他']
    def sort_key(r):
        return (CATEGORY_ORDER.index(r['品类']) if r['品类'] in CATEGORY_ORDER else 99,
                r['系列'], r['SKU名称'])
    skus_sorted = sorted(skus, key=sort_key)

    # 统计
    total_qty   = sum(s['当前数量'] for s in skus_sorted)
    priced      = [s for s in skus_sorted if price_map.get(s['SKU名称'], 0) > 0]
    unpriced    = [s for s in skus_sorted if price_map.get(s['SKU名称'], 0) == 0]
    cat_counts  = {}
    for s in skus_sorted:
        cat_counts[s['品类']] = cat_counts.get(s['品类'], 0) + 1

    # ── 按品类分组生成表格行
    from itertools import groupby
    sections_html = ""
    for cat, grp in groupby(skus_sorted, key=lambda x: x['品类']):
        items = list(grp)
        cat_qty = sum(i['当前数量'] for i in items)
        rows_html = ""
        for i, item in enumerate(items):
            price = price_map.get(item['SKU名称'], 0)
            price_cell = f'<td class="num alert">{price:,}</td>' if price > 0 else '<td class="num muted">—</td>'
            note_cell  = f'<td class="muted">{item["备注"]}</td>' if item['备注'] else '<td></td>'
            rows_html += f"""
      <tr>
        <td>{item['SKU名称']}</td>
        <td>{item['系列']}</td>
        <td class="num">{item['当前数量']}</td>
        {price_cell}
        {note_cell}
      </tr>"""

        sections_html += f"""
<div class="section">
  <div class="section-title">{cat} <span>{len(items)} 个 SKU · 共 {cat_qty} 件</span></div>
  <table>
    <thead><tr>
      <th>SKU 名称</th>
      <th>系列</th>
      <th class="num">现货数量（件）</th>
      <th class="num">建议清理价（¥）</th>
      <th>备注</th>
    </tr></thead>
    <tbody>{rows_html}
    </tbody>
    <tfoot><tr>
      <td colspan="2">小计</td>
      <td class="num">{cat_qty}</td>
      <td colspan="2"></td>
    </tr></tfoot>
  </table>
</div>
"""

    # 未标价提示
    unpriced_tip = ""
    if unpriced:
        names = "、".join(s['SKU名称'] for s in unpriced[:5])
        more  = f"等共 {len(unpriced)} 个" if len(unpriced) > 5 else ""
        unpriced_tip = f"""
<div class="alert-box" style="margin-top:24px">
  <div class="alert-box-title">⚑ 以下 {len(unpriced)} 个 SKU 暂未标价</div>
  <p style="font-size:10px;color:#555555;line-height:1.8">{names}{more}，价格待确认后更新。</p>
</div>"""

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>清理库存报价表 · 宇宙销售统计中心</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#FAF9F6;color:#111111;font-family:'PingFang SC','Microsoft YaHei',Arial,sans-serif;font-size:12px;line-height:1.6}}
.page{{max-width:960px;margin:0 auto;padding:40px 48px}}

.header{{display:flex;justify-content:space-between;align-items:flex-end;padding-bottom:12px;border-bottom:1px solid #111111;margin-bottom:32px}}
.header-left h1{{font-family:Georgia,'Noto Serif SC',serif;font-size:22px;font-weight:700;letter-spacing:0.3px}}
.header-left p{{font-size:10px;color:#555555;margin-top:4px}}
.header-right{{font-size:10px;color:#555555;text-align:right}}
.burgundy-line{{height:2px;background:#8B1A1A;width:48px;margin-top:8px}}

.kpi-row{{display:grid;grid-template-columns:repeat(3,1fr);gap:0;border:1px solid #E8E8E8;margin-bottom:32px}}
.kpi{{padding:20px 24px;border-right:1px solid #E8E8E8}}
.kpi:last-child{{border-right:none}}
.kpi-label{{font-size:9px;color:#555555;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:8px}}
.kpi-value{{font-family:'Calibri',Arial,sans-serif;font-size:28px;font-weight:700;color:#111111;font-variant-numeric:tabular-nums}}
.kpi-sub{{font-size:9px;color:#555555;margin-top:4px}}

.section{{margin-bottom:28px}}
.section-title{{font-family:Georgia,'Noto Serif SC',serif;font-size:15px;font-weight:600;color:#111111;padding-bottom:8px;margin-bottom:14px;border-bottom:1px solid #CCCCCC;display:flex;align-items:center;gap:12px}}
.section-title span{{font-size:9px;color:#555555;font-family:Arial,sans-serif;font-weight:400;text-transform:uppercase;letter-spacing:0.8px}}

table{{width:100%;border-collapse:collapse;font-size:11px}}
thead tr{{background:#111111}}
thead th{{color:#FFFFFF;font-weight:600;font-size:10px;padding:8px 12px;text-align:left}}
thead th.num{{text-align:right}}
tbody tr:nth-child(even){{background:#F5F5F5}}
tbody tr:nth-child(odd){{background:#FFFFFF}}
tbody td{{padding:7px 12px;color:#111111;border-bottom:1px solid #CCCCCC}}
tbody td.num{{text-align:right;font-family:'Calibri',Arial,sans-serif;font-variant-numeric:tabular-nums}}
tbody td.alert{{color:#8B1A1A;font-weight:600;text-align:right}}
tbody td.muted{{color:#555555;text-align:right}}
tfoot tr{{background:#E8E8E8;border-top:2px solid #111111}}
tfoot td{{padding:8px 12px;font-weight:700;font-size:10px}}
tfoot td.num{{text-align:right;font-family:'Calibri',Arial,sans-serif}}

.alert-box{{border:1px solid #CCCCCC;padding:16px 20px;background:#FFFFFF;margin-bottom:28px}}
.alert-box-title{{font-size:9px;font-weight:700;color:#8B1A1A;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:12px}}

.disclaimer{{font-size:9px;color:#555555;background:#F5F5F5;border:1px solid #E8E8E8;padding:12px 16px;margin:28px 0;line-height:1.8}}
.source{{font-size:9px;color:#888888;margin-top:32px;padding-top:10px;border-top:1px solid #CCCCCC}}
</style>
</head>
<body>
<div class="page">

<div class="header">
  <div class="header-left">
    <h1>清理库存 · 渠道报价表</h1>
    <p>宇宙销售统计中心 · 数据截至 2026年{date_disp}</p>
    <div class="burgundy-line"></div>
  </div>
  <div class="header-right">
    对外报价版本<br>生成时间：{now_str}
  </div>
</div>

<div class="kpi-row">
  <div class="kpi">
    <div class="kpi-label">清理 SKU 总数</div>
    <div class="kpi-value">{len(skus_sorted)}</div>
    <div class="kpi-sub">含 {len(priced)} 个已标价 · {len(unpriced)} 个待确认</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">合计现货数量</div>
    <div class="kpi-value">{total_qty}</div>
    <div class="kpi-sub">件，均为公路车品类</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">品类数</div>
    <div class="kpi-value">{len(cat_counts)}</div>
    <div class="kpi-sub">{'、'.join(list(cat_counts.keys())[:4])}{'等' if len(cat_counts)>4 else ''}</div>
  </div>
</div>

<div class="disclaimer">
  <strong>说明：</strong>以上价格为建议清理价，实际成交价以双方书面确认为准。库存数量实时变动，请在正式下单前再次确认现货。本文件仅供授权合作伙伴参考，请勿对外传播。
</div>

{sections_html}

{unpriced_tip}

<div class="source">数据来源：内部库存管理系统 {date_disp} 快照 · 仅含清理级别商品 · 不含成本信息</div>

</div>
</body>
</html>"""

    os.makedirs("库存报表", exist_ok=True)
    with open(OUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ HTML 已保存：{OUT_HTML}")


# ── 内部版 HTML（含成本、价差）────────────────────────────
def build_html_internal(skus, sku_data):
    ts = re.search(r'(\d{4})(\d{2})(\d{2})_v', SOURCE_FILE)
    date_disp = f"{int(ts.group(2))}月{int(ts.group(3))}日" if ts else "3月13日"
    now_str   = datetime.now().strftime("%Y-%m-%d %H:%M")

    CATEGORY_ORDER = ['整车','车架','套件组','牙盘组','轮组','外胎','座垫','锁踏',
                      '碟片组','车把一体把','头盔','眼镜','骑行服','其他']
    def sort_key(r):
        return (CATEGORY_ORDER.index(r['品类']) if r['品类'] in CATEGORY_ORDER else 99,
                r['系列'], r['SKU名称'])
    skus_sorted = sorted(skus, key=sort_key)

    # 汇总 KPI
    total_qty      = sum(s['当前数量'] for s in skus_sorted)
    total_cost     = sum(s['当前数量'] * sku_data.get(s['SKU名称'], {}).get('成本价', 0) for s in skus_sorted)
    total_price    = sum(s['当前数量'] * sku_data.get(s['SKU名称'], {}).get('清理底价', 0)
                        for s in skus_sorted if sku_data.get(s['SKU名称'], {}).get('清理底价', 0) > 0)
    priced_skus    = [s for s in skus_sorted if sku_data.get(s['SKU名称'], {}).get('清理底价', 0) > 0]
    unpriced_skus  = [s for s in skus_sorted if sku_data.get(s['SKU名称'], {}).get('清理底价', 0) == 0]

    from itertools import groupby
    sections_html = ""
    for cat, grp in groupby(skus_sorted, key=lambda x: x['品类']):
        items     = list(grp)
        cat_qty   = sum(i['当前数量'] for i in items)
        cat_cost  = sum(i['当前数量'] * sku_data.get(i['SKU名称'], {}).get('成本价', 0) for i in items)
        cat_price = sum(i['当前数量'] * sku_data.get(i['SKU名称'], {}).get('清理底价', 0)
                       for i in items if sku_data.get(i['SKU名称'], {}).get('清理底价', 0) > 0)

        rows_html = ""
        for i, item in enumerate(items):
            d         = sku_data.get(item['SKU名称'], {})
            cost      = d.get('成本价', 0)
            clr_price = d.get('清理底价', 0)
            qty       = item['当前数量']
            total_c   = qty * cost
            total_p   = qty * clr_price if clr_price > 0 else 0
            diff      = clr_price - cost if (clr_price > 0 and cost > 0) else None
            diff_rate = diff / cost if (diff is not None and cost > 0) else None

            cost_cell  = f'<td class="num">{cost:,.0f}</td>'          if cost > 0      else '<td class="num muted">—</td>'
            tc_cell    = f'<td class="num">{total_c:,.0f}</td>'        if total_c > 0   else '<td class="num muted">—</td>'
            price_cell = f'<td class="num alert">{clr_price:,.0f}</td>' if clr_price > 0 else '<td class="num muted">待定</td>'
            tp_cell    = f'<td class="num alert">{total_p:,.0f}</td>'   if total_p > 0   else '<td class="num muted">—</td>'

            if diff_rate is not None:
                rate_cls  = "pos" if diff_rate >= 0 else "neg"
                diff_cell = f'<td class="num {rate_cls}">{diff_rate:+.1%}</td>'
            else:
                diff_cell = '<td class="num muted">—</td>'

            rows_html += f"""
      <tr>
        <td>{item['SKU名称']}</td>
        <td>{item['系列']}</td>
        <td class="num">{qty}</td>
        {cost_cell}
        {tc_cell}
        {price_cell}
        {tp_cell}
        {diff_cell}
      </tr>"""

        # 小计行
        diff_rate_cat = (cat_price - cat_cost) / cat_cost if cat_cost > 0 else None
        dr_str = f"{diff_rate_cat:+.1%}" if diff_rate_cat is not None else "—"
        sections_html += f"""
<div class="section">
  <div class="section-title">{cat} <span>{len(items)} 个 SKU · 共 {cat_qty} 件</span></div>
  <table>
    <thead><tr>
      <th>SKU 名称</th><th>系列</th>
      <th class="num">数量</th>
      <th class="num">成本价（¥）</th><th class="num">库存成本（¥）</th>
      <th class="num">清理底价（¥）</th><th class="num">清理总价（¥）</th>
      <th class="num">价差率</th>
    </tr></thead>
    <tbody>{rows_html}
    </tbody>
    <tfoot><tr>
      <td colspan="2">小计</td>
      <td class="num">{cat_qty}</td>
      <td></td>
      <td class="num">{cat_cost:,.0f}</td>
      <td></td>
      <td class="num">{cat_price:,.0f}</td>
      <td class="num">{dr_str}</td>
    </tr></tfoot>
  </table>
</div>
"""

    # 未标价提示
    unpriced_tip = ""
    if unpriced_skus:
        names = "、".join(s['SKU名称'] for s in unpriced_skus[:5])
        more  = f"等共 {len(unpriced_skus)} 个" if len(unpriced_skus) > 5 else ""
        unpriced_tip = f"""
<div class="alert-box">
  <div class="alert-box-title">⚑ {len(unpriced_skus)} 个 SKU 暂未标清理底价</div>
  <p style="font-size:10px;color:#555555;line-height:1.8">{names}{more}，其库存成本仍计入合计，但清理总价列显示为空。</p>
</div>"""

    overall_rate = (total_price - total_cost) / total_cost if total_cost > 0 else 0

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>清理库存内部版 · 含成本 · 宇宙销售统计中心</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#FAF9F6;color:#111111;font-family:'PingFang SC','Microsoft YaHei',Arial,sans-serif;font-size:12px;line-height:1.6}}
.page{{max-width:1100px;margin:0 auto;padding:40px 48px}}

.header{{display:flex;justify-content:space-between;align-items:flex-end;padding-bottom:12px;border-bottom:1px solid #111111;margin-bottom:32px}}
.header-left h1{{font-family:Georgia,'Noto Serif SC',serif;font-size:22px;font-weight:700;letter-spacing:0.3px}}
.header-left p{{font-size:10px;color:#555555;margin-top:4px}}
.header-right{{font-size:10px;color:#555555;text-align:right;line-height:1.8}}
.burgundy-line{{height:2px;background:#8B1A1A;width:48px;margin-top:8px}}
.badge-internal{{display:inline-block;background:#8B1A1A;color:#fff;font-size:9px;font-weight:700;letter-spacing:1px;padding:2px 8px;margin-left:10px;vertical-align:middle}}

.kpi-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:0;border:1px solid #E8E8E8;margin-bottom:32px}}
.kpi{{padding:18px 22px;border-right:1px solid #E8E8E8}}
.kpi:last-child{{border-right:none}}
.kpi-label{{font-size:9px;color:#555555;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:8px}}
.kpi-value{{font-family:'Calibri',Arial,sans-serif;font-size:26px;font-weight:700;color:#111111;font-variant-numeric:tabular-nums}}
.kpi-value.alert{{color:#8B1A1A}}
.kpi-sub{{font-size:9px;color:#555555;margin-top:4px}}

.section{{margin-bottom:28px}}
.section-title{{font-family:Georgia,'Noto Serif SC',serif;font-size:15px;font-weight:600;color:#111111;padding-bottom:8px;margin-bottom:14px;border-bottom:1px solid #CCCCCC;display:flex;align-items:center;gap:12px}}
.section-title span{{font-size:9px;color:#555555;font-family:Arial,sans-serif;font-weight:400;text-transform:uppercase;letter-spacing:0.8px}}

table{{width:100%;border-collapse:collapse;font-size:11px}}
thead tr{{background:#111111}}
thead th{{color:#FFFFFF;font-weight:600;font-size:10px;padding:8px 12px;text-align:left}}
thead th.num{{text-align:right}}
tbody tr:nth-child(even){{background:#F5F5F5}}
tbody tr:nth-child(odd){{background:#FFFFFF}}
tbody td{{padding:7px 12px;color:#111111;border-bottom:1px solid #CCCCCC}}
tbody td.num{{text-align:right;font-family:'Calibri',Arial,sans-serif;font-variant-numeric:tabular-nums}}
tbody td.alert{{color:#8B1A1A;font-weight:600;text-align:right}}
tbody td.muted{{color:#888888;text-align:right}}
tbody td.pos{{color:#2a6041;font-weight:600;text-align:right}}
tbody td.neg{{color:#8B1A1A;font-weight:600;text-align:right}}
tfoot tr{{background:#E8E8E8;border-top:2px solid #111111}}
tfoot td{{padding:8px 12px;font-weight:700;font-size:10px}}
tfoot td.num{{text-align:right;font-family:'Calibri',Arial,sans-serif}}

.alert-box{{border:1px solid #CCCCCC;padding:16px 20px;background:#FFFFFF;margin-bottom:28px}}
.alert-box-title{{font-size:9px;font-weight:700;color:#8B1A1A;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:8px}}
.source{{font-size:9px;color:#888888;margin-top:32px;padding-top:10px;border-top:1px solid #CCCCCC}}
</style>
</head>
<body>
<div class="page">

<div class="header">
  <div class="header-left">
    <h1>清理库存 · 成本核查表 <span class="badge-internal">INTERNAL</span></h1>
    <p>宇宙销售统计中心 · 数据截至 2026年{date_disp}</p>
    <div class="burgundy-line"></div>
  </div>
  <div class="header-right">
    内部文件 · 含成本信息 · 请勿外传<br>生成时间：{now_str}
  </div>
</div>

<div class="kpi-row">
  <div class="kpi">
    <div class="kpi-label">清理 SKU 总数</div>
    <div class="kpi-value">{len(skus_sorted)}</div>
    <div class="kpi-sub">有库存 · 清理 + Pinarello 清理</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">库存成本合计</div>
    <div class="kpi-value alert">¥{total_cost/10000:.1f}万</div>
    <div class="kpi-sub">¥{total_cost:,.0f}</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">已标价清理总额</div>
    <div class="kpi-value">¥{total_price/10000:.1f}万</div>
    <div class="kpi-sub">¥{total_price:,.0f} · {len(priced_skus)} 个已标价</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">整体价差率</div>
    <div class="kpi-value {'alert' if overall_rate < 0 else ''}">{overall_rate:+.1%}</div>
    <div class="kpi-sub">{'低于成本，需关注' if overall_rate < 0 else '高于成本'}</div>
  </div>
</div>

{unpriced_tip}

{sections_html}

<div class="source">数据来源：公路车库存报表 {date_disp} 版本 · 含成本信息 · 仅限内部使用</div>
</div>
</body>
</html>"""

    with open(OUT_HTML_INTERNAL, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ 内部版 HTML 已保存：{OUT_HTML_INTERNAL}")
    print(f"   库存成本合计：¥{total_cost:,.0f}  /  已标价清理总额：¥{total_price:,.0f}  /  整体价差率：{overall_rate:+.1%}")


# ── 主程序 ─────────────────────────────────────────────────
def main():
    print(f"读取：{SOURCE_FILE}")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    sku_data  = load_sku_data(wb)
    price_map = {k: v['清理底价'] for k, v in sku_data.items()}
    print(f"  SKU 数据字典：{len(sku_data)} 个")

    skus = load_clearance_skus(wb)
    print(f"  清理 SKU（有库存）：{len(skus)} 个")

    build_excel(skus, price_map)
    build_html(skus, price_map)
    build_html_internal(skus, sku_data)

if __name__ == "__main__":
    main()
